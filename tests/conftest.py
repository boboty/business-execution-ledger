from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from sqlalchemy.orm import Session

from bel.infrastructure.persistence.database import make_engine, make_session_factory
from bel.infrastructure.persistence.models import Base
from tests.pg_disposable import DISPOSABLE_SKIP_REASON, resolve_disposable_postgres_url

PRIMARY_SHEET_NAME = "报关出口购销合同"


def pytest_collection_modifyitems(config: pytest.Config, items: list[pytest.Item]) -> None:
    """Auto-skip everything marked ``@pytest.mark.postgres`` unless the
    disposable test-database contract is satisfied — see the ``postgres``
    marker's registration in pyproject.toml, ``tests/pg_disposable.py``
    for the contract, and the ``postgres_url`` fixture below for the
    fixture-level double guard. Keeps plain local `pytest` runs passing
    with zero PostgreSQL dependency (Part F: SQLite test convenience vs.
    PostgreSQL production contract, kept separate), and — critically —
    means the destructive PostgreSQL fixtures can never activate from
    ``BEL_DATABASE_URL`` alone: the runtime database is never dropped or
    reset by the test suite."""
    if resolve_disposable_postgres_url(os.environ) is not None:
        return
    skip_postgres = pytest.mark.skip(reason=DISPOSABLE_SKIP_REASON)
    for item in items:
        if "postgres" in item.keywords:
            item.add_marker(skip_postgres)


@pytest.fixture
def postgres_url() -> str:
    """The disposable PostgreSQL test database URL — never the runtime
    ``BEL_DATABASE_URL``. Second, fixture-level guard behind the
    collection-level skip above: if a test somehow reaches fixture setup
    without the disposable contract being satisfied, skip here — BEFORE
    the test body (and its DROPs) can run — rather than ever handing out
    a runtime database."""
    url = resolve_disposable_postgres_url(os.environ)
    if url is None:
        pytest.skip(DISPOSABLE_SKIP_REASON)
    return url


@pytest.fixture
def db_session() -> Session:
    """In-memory SQLite with schema created directly from the ORM models.
    Schema-shape fidelity against Alembic migrations is covered
    separately by tests/integration/test_migration.py."""
    engine = make_engine("sqlite://")
    Base.metadata.create_all(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as session:
        yield session


@pytest.fixture
def phase2b_ledger_path(ledger_workbook_factory) -> Path:
    from fixtures.synthetic.phase2b_close import PHASE2B_CONTRACT_HEADERS, PHASE2B_CONTRACT_ROWS

    return ledger_workbook_factory(
        PHASE2B_CONTRACT_HEADERS, PHASE2B_CONTRACT_ROWS, filename="phase2b-contracts.xlsx"
    )


@pytest.fixture
def phase2b_invoices_path(invoice_workbook_factory) -> Path:
    from fixtures.synthetic import scenarios
    from fixtures.synthetic.phase2b_close import PHASE2B_INVOICE_ROWS

    return invoice_workbook_factory(PHASE2B_INVOICE_ROWS, buyer=scenarios.BUYER, filename="phase2b-invoices.xlsx")


@pytest.fixture
def phase2b_close_facts_path(tmp_path) -> Path:
    from fixtures.synthetic.phase2b_close import write_phase2b_close_facts

    return write_phase2b_close_facts(tmp_path / "phase2b-close-facts.json")


@pytest.fixture
def phase2b_recompute_facts_path(tmp_path) -> Path:
    from fixtures.synthetic.phase2b_close import write_recompute_fact_pack

    return write_recompute_fact_pack(tmp_path / "phase2b-recompute-facts.json")


def write_ledger_workbook(
    path: Path,
    headers: list[str],
    data_rows: list[list[Any]],
    sheet_name: str = PRIMARY_SHEET_NAME,
    extra_sheets: list[str] | None = None,
) -> None:
    """Build a minimal workbook matching the adapter's documented shape: row 1 is
    a title row (ignored), row 2 is headers, data starts at row 3."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Synthetic Fixture Title Row"])
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    for name in extra_sheets or []:
        wb.create_sheet(title=name)
    wb.save(path)


@pytest.fixture
def ledger_workbook_factory(tmp_path: Path):
    def _factory(headers, data_rows, sheet_name=PRIMARY_SHEET_NAME, extra_sheets=None, filename="ledger.xlsx"):
        path = tmp_path / filename
        write_ledger_workbook(path, headers, data_rows, sheet_name=sheet_name, extra_sheets=extra_sheets)
        return path

    return _factory


INVOICE_SHEET_NAME = "sheet1"
INVOICE_HEADERS = [
    "凭证模板",
    "凭证字号",
    "发票票种",
    "开票日期",
    "发票号码",
    "数电发票号码",
    "销方名称",
    "商品名称（明细）",
    "规格型号（明细）",
    "单位（明细）",
    "数量（明细）",
    "单价（明细）",
    "金额（明细）",
    "税率（%）（明细）",
    "税额（明细）",
    "价税合计（明细）",
    "发票状态",
    "发票金额",
    "发票税额",
    "发票价税合计",
]


def write_invoice_workbook(path: Path, buyer: str, data_rows: list[list[Any]]) -> None:
    """Row 1 blank, row 2 col A is the buyer, row 3 is headers, data
    starts row 4 — matches the real invoice ledger's shape."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = INVOICE_SHEET_NAME
    ws.append([])
    ws.append([buyer])
    ws.append(INVOICE_HEADERS)
    for row in data_rows:
        ws.append(row)
    wb.save(path)


@pytest.fixture
def invoice_workbook_factory(tmp_path: Path):
    def _factory(data_rows, buyer="Buyer Co", filename="invoices.xlsx"):
        path = tmp_path / filename
        write_invoice_workbook(path, buyer, data_rows)
        return path

    return _factory

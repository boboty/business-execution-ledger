"""Phase 2D.4-F2 — Exception & Task Data Product over a real session.

DB-seeded scenarios proving the Data Product is a faithful, lossless
transformation of the F1 Center: a genuine mixed source set (TASK_EXCEPTION
+ MATCH_CASE + period COMPUTED_BLOCKER), unmappable tasks retained,
multi-scope id sets complete, computed-blocker semantics (created_at blank,
F1 deterministic source_id, period preserved), no advisory /
MISSING_CONTRACT_GROSS_AMOUNT / UNMATCHED, and filter parity — the page
projection's item identity set equals the CSV and XLSX identity sets for
every meaningful filter.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal

import openpyxl
import pytest

from bel.application.exception_task_data_product import (
    build_exception_task_data_product,
    export_exception_task_csv,
    export_exception_task_xlsx,
)
from bel.application.sales_contract_facts import create_sales_contract_fact
from bel.application.unresolved_work_center import (
    SourceType,
    UnresolvedWorkFilters,
    get_unresolved_work_center,
)
from bel.domain.accrual import CostRecognitionFact
from bel.domain.contract import Contract
from bel.domain.evidence import EvidenceDocument, EvidenceFragment, FragmentKind
from bel.domain.exception import ExceptionStatus, ExceptionType, TaskException
from bel.domain.invoice import Invoice, InvoiceDirection
from bel.domain.matching import (
    MatchCase,
    MatchCaseStatus,
    MatchCandidate,
    MatchMethod,
    SubjectType,
)
from bel.infrastructure.persistence.database import make_engine, make_session_factory
from bel.infrastructure.persistence.models import Base
from bel.infrastructure.persistence.repositories import (
    ContractRepository,
    CostRecognitionFactRepository,
    EvidenceRepository,
    ExceptionRepository,
    InvoiceRepository,
    MatchCandidateRepository,
    MatchCaseRepository,
)

NOW = datetime.now(timezone.utc)


@pytest.fixture
def db_session():
    engine = make_engine("sqlite://")
    Base.metadata.create_all(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as session:
        yield session


def _make_fragment(session):
    doc = EvidenceDocument(
        id=uuid.uuid4(), file_name="x", sha256=uuid.uuid4().hex + uuid.uuid4().hex, source_type="t", imported_at=NOW
    )
    EvidenceRepository(session).add_document(doc)
    frag = EvidenceFragment(
        id=uuid.uuid4(),
        evidence_document_id=doc.id,
        fragment_kind=FragmentKind.MANUAL_FACT,
        sheet_name=None,
        row_number=None,
        locator_json={"index": 0},
        raw_data={},
        created_at=NOW,
    )
    EvidenceRepository(session).add_fragment(frag)
    session.flush()
    return frag


def _make_contract(session, fragment_id, contract_no=None):
    contract = Contract(
        id=uuid.uuid4(),
        contract_no=contract_no or f"C-{uuid.uuid4().hex[:8]}",
        contract_type=None,
        counterparty="Supplier A",
        buyer="Our Own Entity",
        gross_amount=Decimal("5000.00"),
        currency="CNY",
        contract_date=date(2026, 1, 1),
        current_source_fragment_id=fragment_id,
        created_at=NOW,
        updated_at=NOW,
    )
    ContractRepository(session).add(contract)
    session.flush()
    return contract


def _make_task(session, exception_type, detail, summary="task"):
    task = TaskException(
        id=uuid.uuid4(),
        exception_type=exception_type,
        status=ExceptionStatus.OPEN,
        summary=summary,
        detail=detail,
        created_at=NOW,
    )
    ExceptionRepository(session).add(task)
    session.flush()
    return task


def _make_invoice(session, fragment_id, external_key):
    invoice = Invoice(
        id=uuid.uuid4(),
        direction=InvoiceDirection.PURCHASE,
        invoice_type=None,
        invoice_no=None,
        digital_invoice_no=None,
        external_invoice_key=external_key,
        issue_date=date(2026, 1, 5),
        seller="Seller Co",
        buyer="Buyer Co",
        net_amount=Decimal("100.00"),
        tax_amount=Decimal("0"),
        gross_amount=Decimal("100.00"),
        invoice_status=None,
        source_fragment_id=fragment_id,
        created_at=NOW,
        updated_at=NOW,
    )
    InvoiceRepository(session).add(invoice)
    session.flush()
    return invoice


def _seed_mixed_center(session):
    """A genuine mixed source set: two contracts, a multi-scope task, an
    unmappable task, a procurement HCR MatchCase, and a blocker-producing
    contract."""
    frag = _make_fragment(session)
    c1 = _make_contract(session, frag.id, "PO-DP-A")
    c2 = _make_contract(session, frag.id, "PO-DP-B")

    multi_task = _make_task(
        session,
        ExceptionType.BUSINESS_KEY_CONFLICT,
        {"contract_ids": [str(c1.id), str(c2.id)]},
        summary="采购合同编号冲突",
    )
    unmappable = _make_task(
        session,
        ExceptionType.SALES_CONTRACT_IDENTITY_INCOMPLETE,
        {"source_fragment_id": str(uuid.uuid4())},
        summary="外销合同身份信息不完整",
    )

    invoice = _make_invoice(session, frag.id, "INV-DP-1")
    match_case = MatchCase(
        id=uuid.uuid4(),
        subject_type=SubjectType.INVOICE,
        subject_id=invoice.id,
        status=MatchCaseStatus.HUMAN_CONFIRMATION_REQUIRED,
        match_method=MatchMethod.M001,
        created_at=NOW,
        resolved_at=None,
    )
    MatchCaseRepository(session).add(match_case)
    session.flush()
    MatchCandidateRepository(session).add(
        MatchCandidate(id=uuid.uuid4(), match_case_id=match_case.id, contract_id=c1.id, created_at=NOW)
    )
    session.flush()

    # MISSING_ACCRUAL_BASIS blocker for any later period.
    CostRecognitionFactRepository(session).add(
        CostRecognitionFact(
            id=uuid.uuid4(),
            contract_id=c1.id,
            recognition_date=date(2026, 2, 28),
            basis="MANUAL_CONFIRMED",
            source_fragment_id=frag.id,
            created_at=NOW,
        )
    )
    session.flush()
    session.commit()
    return {
        "multi_task": multi_task,
        "unmappable": unmappable,
        "match_case": match_case,
        "contract_a": c1,
        "contract_b": c2,
    }


def _identity_set(rows) -> set[tuple[str, str]]:
    return {(r["record_type"], r["source_id"]) for r in rows}


def _csv_rows(product) -> list[dict]:
    text = export_exception_task_csv(product).decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _xlsx_rows(product) -> set[tuple[str, str]]:
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    result = set()
    for sheet, source_type in (
        ("02_System_Tasks", SourceType.TASK_EXCEPTION),
        ("03_Match_Confirmation", SourceType.MATCH_CASE),
        ("04_Period_Close_Blockers", SourceType.COMPUTED_BLOCKER),
    ):
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_map = dict(zip(header, row))
            result.add((row_map["source_type"], row_map["source_id"]))
    return result


def test_mixed_scenario_shape_and_counts(db_session):
    seeded = _seed_mixed_center(db_session)
    center = get_unresolved_work_center(db_session, filters=UnresolvedWorkFilters(period="2026-03"))
    product = build_exception_task_data_product(center)

    assert center.counts[SourceType.TASK_EXCEPTION] == 2
    assert center.counts[SourceType.MATCH_CASE] == 1
    assert center.counts[SourceType.COMPUTED_BLOCKER] >= 1

    # Four exact XLSX sheets.
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    assert wb.sheetnames == [
        "01_Summary",
        "02_System_Tasks",
        "03_Match_Confirmation",
        "04_Period_Close_Blockers",
    ]

    # Unified CSV record_type.
    rows = _csv_rows(product)
    assert {r["record_type"] for r in rows} == {
        SourceType.TASK_EXCEPTION,
        SourceType.MATCH_CASE,
        SourceType.COMPUTED_BLOCKER,
    }

    # Canonical source identity survives byte-for-byte.
    task_rows = [r for r in rows if r["record_type"] == SourceType.TASK_EXCEPTION]
    assert {r["source_id"] for r in task_rows} == {str(seeded["multi_task"].id), str(seeded["unmappable"].id)}
    match_rows = [r for r in rows if r["record_type"] == SourceType.MATCH_CASE]
    assert [r["source_id"] for r in match_rows] == [str(seeded["match_case"].id)]

    # Summary counts + period, no generated_at.
    summary_ws = wb["01_Summary"]
    field_value = {
        summary_ws.cell(row=r, column=1).value: summary_ws.cell(row=r, column=2).value
        for r in range(1, summary_ws.max_row + 1)
    }
    assert field_value["period"] == "2026-03"
    assert field_value["TASK_EXCEPTION"] == 2
    assert field_value["MATCH_CASE"] == 1
    assert field_value["COMPUTED_BLOCKER"] >= 1
    assert "generated_at" not in field_value


def test_multi_scope_ids_complete_via_db(db_session):
    seeded = _seed_mixed_center(db_session)
    center = get_unresolved_work_center(db_session)
    product = build_exception_task_data_product(center)
    row = next(r for r in _csv_rows(product) if r["source_id"] == str(seeded["multi_task"].id))
    ids = json.loads(row["procurement_contract_ids"])
    assert sorted(str(x) for x in (seeded["contract_a"].id, seeded["contract_b"].id)) == ids
    assert len(ids) == 2  # never truncated to a first id


def test_unmappable_task_retained_with_blank_scope(db_session):
    seeded = _seed_mixed_center(db_session)
    center = get_unresolved_work_center(db_session)
    product = build_exception_task_data_product(center)
    row = next(r for r in _csv_rows(product) if r["source_id"] == str(seeded["unmappable"].id))
    assert row["procurement_contract_ids"] == ""
    assert row["sales_contract_ids"] == ""
    assert row["scopes_json"] == ""
    # The task's own id is the identity — never a guessed Contract.
    assert row["source_id"] == str(seeded["unmappable"].id)


def test_computed_blocker_created_at_blank_and_source_id_is_f1_key(db_session):
    _seed_mixed_center(db_session)
    center = get_unresolved_work_center(db_session, filters=UnresolvedWorkFilters(period="2026-03"))
    product = build_exception_task_data_product(center)
    blocker_rows = [r for r in _csv_rows(product) if r["record_type"] == SourceType.COMPUTED_BLOCKER]
    assert blocker_rows
    for row in blocker_rows:
        assert row["created_at"] == ""
        assert row["status"] == "PRESENT"
        assert row["resolution_route"] == "REVIEW_ONLY"
        assert row["provenance"] == "bel.application.period_close"
        # The F1 deterministic key is used as-is — no export-local id.
        assert "|" in row["source_id"]


def test_no_advisory_no_missing_gross_no_unmatched(db_session):
    _seed_mixed_center(db_session)
    center = get_unresolved_work_center(db_session, filters=UnresolvedWorkFilters(period="2026-03"))
    product = build_exception_task_data_product(center)
    codes = {r["code"] for r in _csv_rows(product)}
    assert "MISSING_CONTRACT_GROSS_AMOUNT" not in codes
    assert not (codes & {"SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED", "PURCHASE_INVOICE_AMOUNT_DEVIATION"})
    assert all(r["record_type"] != "UNMATCHED" for r in _csv_rows(product))
    assert all(r["code"] != "UNMATCHED" for r in _csv_rows(product))


# ---------------------------------------------------------------------------
# Filter parity — page projection identity set == CSV == XLSX
# ---------------------------------------------------------------------------


def _assert_parity(db_session, filters):
    center = get_unresolved_work_center(db_session, filters=filters)
    projection_ids = {(i.source_type, str(i.source_id)) for i in center.items}
    product = build_exception_task_data_product(center)
    csv_ids = _identity_set(_csv_rows(product))
    xlsx_ids = _xlsx_rows(product)
    assert projection_ids == csv_ids == xlsx_ids
    # Every CSV/XLSX row is a projection row — nothing extra, nothing lost.
    assert csv_ids <= projection_ids


def test_filter_parity_source_type(db_session):
    _seed_mixed_center(db_session)
    _assert_parity(db_session, UnresolvedWorkFilters(source_type=SourceType.MATCH_CASE))
    _assert_parity(db_session, UnresolvedWorkFilters(source_type=SourceType.TASK_EXCEPTION))


def test_filter_parity_code(db_session):
    _seed_mixed_center(db_session)
    _assert_parity(db_session, UnresolvedWorkFilters(code=ExceptionType.SALES_CONTRACT_IDENTITY_INCOMPLETE))


def test_filter_parity_procurement_scope(db_session):
    seeded = _seed_mixed_center(db_session)
    _assert_parity(db_session, UnresolvedWorkFilters(procurement_contract_id=seeded["contract_a"].id))


def test_filter_parity_sales_scope(db_session):
    session = db_session
    frag = _make_fragment(session)
    sales = create_sales_contract_fact(
        session,
        our_entity="Our Own Entity",
        sales_contract_no="SC-DP-PARITY",
        fields={"customer": "Customer A"},
        source_fragment_id=frag.id,
        created_at=NOW,
    ).sales_contract
    session.flush()
    _make_task(
        session,
        ExceptionType.SALES_CONTRACT_CUSTOMER_UNRESOLVED,
        {"sales_contract_id": str(sales.id)},
        summary="客户待补充",
    )
    session.commit()
    _assert_parity(session, UnresolvedWorkFilters(sales_contract_id=sales.id))


def test_filter_parity_period(db_session):
    _seed_mixed_center(db_session)
    _assert_parity(db_session, UnresolvedWorkFilters(period="2026-03"))
    _assert_parity(db_session, UnresolvedWorkFilters(period=None))

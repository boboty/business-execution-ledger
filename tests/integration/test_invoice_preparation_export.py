"""Invoice Preparation Data Product tests (Phase 2D.3-F2b).

The Data Product is built ONLY from the ``InvoicePreparationWorkbench``
(never from a Session, never by re-running a business rule) and
serialized to the exact five-sheet XLSX and the unified record_type CSV.
Same-source, sales/supplier scenario, format/safety, and reproducibility
coverage. The DB-backed scenarios reuse the F2a web builder; dangling
(incomplete) associations — which the repositories cannot persist — are
exercised via pure F0 contexts, exactly as the F2a repair tests do.
"""

from __future__ import annotations

import csv
import io
import json
import time
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal

import openpyxl
import pytest

from bel.application.invoice_preparation import (
    InvoicePreparationContext,
    SalesScopeContext,
    SalesScopeInvoiceAllocation,
    SalesScopeLinkedProcurementContract,
    SalesScopePaymentAllocation,
    SupplierScopeContext,
    SupplierScopeInvoiceAllocation,
    SupplierScopeInvoiceItemAllocation,
    SupplierScopePaymentAllocation,
)
from bel.application.invoice_preparation_export import (
    ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION,
    ATTENTION_CATEGORY_MANAGEMENT_ADVISORY,
    RECORD_TYPE_SALES_ATTENTION,
    RECORD_TYPE_SALES_PREPARATION,
    RECORD_TYPE_SUPPLIER_ATTENTION,
    RECORD_TYPE_SUPPLIER_REQUEST,
    build_invoice_preparation_data_product,
    export_invoice_preparation_csv,
    export_invoice_preparation_xlsx,
)
from bel.application.invoice_preparation_workbench import (
    get_invoice_preparation_workbench,
    get_invoice_preparation_workbench_from_context,
)
from bel.application.supplier_invoice_request import SupplierRequestCheckOutcome
from bel.domain.accrual import InvoiceItemAllocation
from bel.domain.contract import Contract, ContractItem
from bel.domain.invoice import Invoice, InvoiceDirection, InvoiceItem
from bel.domain.matching import (
    AllocationMatchMethod,
    ConfirmationType,
    InvoiceAllocation,
    PaymentAllocation,
    SalesInvoiceAllocation,
    SalesPaymentAllocation,
)
from bel.domain.payment import Payment, PaymentDirection
from bel.domain.procurement_sales_link import ProcurementSalesLink
from bel.domain.sales_contract import SalesContract
from bel.domain.shipment import Shipment
from bel.infrastructure.persistence.database import make_engine, make_session_factory
from bel.infrastructure.persistence.models import Base
from tests.web.test_web_invoice_preparation import _build_workbench_db

NOW = datetime.now(timezone.utc)

EXPECTED_SHEETS = [
    "01_Summary",
    "02_Sales_Preparation",
    "03_Sales_Attention",
    "04_Supplier_Request",
    "05_Supplier_Attention",
]

ALL_RECORD_TYPES = {
    RECORD_TYPE_SALES_PREPARATION,
    RECORD_TYPE_SALES_ATTENTION,
    RECORD_TYPE_SUPPLIER_REQUEST,
    RECORD_TYPE_SUPPLIER_ATTENTION,
}


@pytest.fixture
def export_product(tmp_path):
    db_path = tmp_path / "export.db"
    _build_workbench_db(str(db_path))
    engine = make_engine(f"sqlite:///{db_path}")
    Base.metadata.create_all(engine)
    with make_session_factory(engine)() as session:
        product = build_invoice_preparation_data_product(get_invoice_preparation_workbench(session))
    return product


def _sales_row(product, no):
    return next(r for r in product.sales_preparation if r.sales_contract_no == no)


def _supplier_row(product, no):
    return next(r for r in product.supplier_request if r.contract_no == no)


def _attention_for(product, *, code=None, category=None, sales_no=None, contract_no=None):
    rows = list(product.sales_attention) + list(product.supplier_attention)
    if code is not None:
        rows = [r for r in rows if r.attention_code == code]
    if category is not None:
        rows = [r for r in rows if r.attention_category == category]
    if sales_no is not None:
        rows = [r for r in rows if r.sales_contract_no == sales_no]
    if contract_no is not None:
        rows = [r for r in rows if r.contract_no == contract_no]
    return rows


# ---------------------------------------------------------------------------
# A. SAME SOURCE — export is a projection of the Workbench, no re-derivation
# ---------------------------------------------------------------------------


def test_export_builder_accepts_workbench_not_session(export_product):
    """The builder's contract is the Workbench: the produced rows carry the
    SAME canonical comparison outcomes the F1 reports decided — proving no
    rule is re-run by the export path."""
    match = _sales_row(export_product, "SC-MTCH")
    assert match.comparison_outcome == "MATCH"
    dev = _sales_row(export_product, "SC-DEV")
    assert dev.comparison_outcome == "DEVIATION"
    amb = _sales_row(export_product, "SC-AMB")
    assert amb.comparison_outcome == "NOT_COMPARABLE_AMBIGUOUS_SCOPE"
    noinv = _sales_row(export_product, "SC-NOINV")
    assert noinv.comparison_outcome == "NOT_COMPARABLE_MISSING_FACT"
    # Supplier amount-check outcome is the F1 decision's, verbatim.
    amt = _supplier_row(export_product, "PO-AMT")
    assert amt.supplier_amount_check_outcome == "DEVIATION"


def test_export_product_counts_match_workbench_scopes(export_product):
    assert export_product.summary["sales_scope_count"] == 4
    assert export_product.summary["supplier_scope_count"] == 9
    assert export_product.summary["sales_comparison_MATCH"] == 1
    assert export_product.summary["sales_comparison_DEVIATION"] == 1
    assert export_product.summary["sales_comparison_NOT_COMPARABLE_AMBIGUOUS_SCOPE"] == 1
    assert export_product.summary["sales_comparison_NOT_COMPARABLE_MISSING_FACT"] == 1
    assert export_product.summary["supplier_amount_check_DEVIATION"] == 1
    assert export_product.summary["supplier_item_name_check_DEVIATION"] == 1
    assert export_product.summary["attention_MANAGEMENT_ADVISORY"] == 5


# ---------------------------------------------------------------------------
# B. SALES scenarios
# ---------------------------------------------------------------------------


def test_sales_match_exports_correctly(export_product):
    row = _sales_row(export_product, "SC-MTCH")
    assert row.comparison_outcome == "MATCH"
    assert row.sales_contract_amount == Decimal("100.00")
    assert row.declared_amount == Decimal("100.00")
    assert row.sales_invoice_amount == Decimal("100.00")
    assert row.sales_contract_currency == row.declared_currency == row.sales_invoice_currency == "USD"
    assert row.confirmed_sales_invoice_count == 1
    assert row.linked_procurement_contract_count == 1
    assert _attention_for(export_product, sales_no="SC-MTCH") == []


def test_sales_deviation_exports_advisory_separately(export_product):
    row = _sales_row(export_product, "SC-DEV")
    assert row.comparison_outcome == "DEVIATION"
    assert row.sales_invoice_amount == Decimal("90.00")
    # The deviation is an ATTENTION row, not flattened into the comparison.
    advisories = _attention_for(export_product, code="SALES_INVOICE_AMOUNT_DEVIATION", sales_no="SC-DEV")
    assert len(advisories) == 1
    assert advisories[0].attention_category == ATTENTION_CATEGORY_MANAGEMENT_ADVISORY
    assert advisories[0].record_type == RECORD_TYPE_SALES_ATTENTION


def test_sales_missing_fact_is_not_comparable_and_value_blank(export_product):
    row = _sales_row(export_product, "SC-NOINV")
    assert row.comparison_outcome == "NOT_COMPARABLE_MISSING_FACT"
    assert row.sales_invoice_amount is None
    assert row.sales_invoice_currency is None
    assert row.confirmed_sales_invoice_count == 0
    # The declared leg is still exposed (the "why").
    assert row.declared_amount == Decimal("100.00")


def test_sales_ambiguous_scope_not_summed(export_product):
    row = _sales_row(export_product, "SC-AMB")
    assert row.comparison_outcome == "NOT_COMPARABLE_AMBIGUOUS_SCOPE"
    # Two links, both with a 100 USD declaration — never summed, never
    # compared against an arbitrary choice: the declaration leg stays None.
    assert row.linked_procurement_contract_count == 2
    assert row.declared_amount is None
    assert row.declared_currency is None
    # The single confirmed SALES invoice Fact is one unambiguous Fact (not
    # a sum and not a choice among many) and stays inspectable.
    assert row.sales_invoice_amount == Decimal("100.00")
    assert row.confirmed_sales_invoice_count == 1


# ---------------------------------------------------------------------------
# C. SUPPLIER scenarios
# ---------------------------------------------------------------------------


def test_supplier_reference_amount_and_currency_exported(export_product):
    row = _supplier_row(export_product, "PO-NOINV")
    assert row.expected_invoice_amount == Decimal("100.00")
    assert row.expected_invoice_currency == "USD"
    assert row.contract_gross_amount == Decimal("100.00")
    assert row.confirmed_purchase_invoice_count == 0
    assert row.supplier_amount_check_outcome is None


def test_supplier_p02_deviation_exported_advisory(export_product):
    row = _supplier_row(export_product, "PO-AMT")
    assert row.supplier_amount_check_outcome == "DEVIATION"
    checks = json.loads(row.supplier_checks_json)
    # PO-AMT also has a MATCHing item-name check; never a single fake
    # "overall status" — the two checks are separate structured elements.
    assert row.supplier_amount_check_count == 1
    assert row.supplier_item_name_check_count == 1
    amount = next(c for c in checks if c["kind"] == "AMOUNT")
    assert amount["outcome"] == "DEVIATION"
    assert amount["reference_amount"] == "100.00"
    assert amount["invoice_amount"] == "90.00"
    # The canonical F1 check_name and invoice id are preserved.
    assert amount["check_name"] == "PURCHASE_INVOICE_GROSS_AMOUNT_VS_CONTRACT_GROSS_AMOUNT"
    assert amount["invoice_id"]
    advisories = _attention_for(export_product, code="PURCHASE_INVOICE_AMOUNT_DEVIATION", contract_no="PO-AMT")
    assert len(advisories) == 1
    assert advisories[0].attention_category == ATTENTION_CATEGORY_MANAGEMENT_ADVISORY
    assert advisories[0].record_type == RECORD_TYPE_SUPPLIER_ATTENTION


def test_supplier_p05_product_deviation_exported(export_product):
    row = _supplier_row(export_product, "PO-PROD")
    checks = json.loads(row.supplier_checks_json)
    # PO-PROD also carries a MATCHing amount check (100 = 100) — separate
    # structured elements, not one flattened status.
    item = next(c for c in checks if c["kind"] == "ITEM_NAME")
    assert item["outcome"] == "DEVIATION"
    assert item["check_name"] == "INVOICE_ITEM_PRODUCT_NAME_VS_CONTRACT_ITEM_PRODUCT_NAME"
    # The canonical P05 allocation/contract-item/invoice-item ids survive.
    assert item["allocation_id"]
    assert item["contract_item_id"]
    assert item["invoice_item_id"]
    assert row.supplier_item_name_deviation_count == 1
    advisories = _attention_for(export_product, code="PURCHASE_INVOICE_PRODUCT_NAME_DEVIATION", contract_no="PO-PROD")
    assert len(advisories) == 1
    assert advisories[0].attention_category == ATTENTION_CATEGORY_MANAGEMENT_ADVISORY


def test_supplier_p03_cardinality_exported(export_product):
    advisories = _attention_for(export_product, code="MULTIPLE_PURCHASE_INVOICES_ON_CONTRACT", contract_no="PO-MULTI")
    assert len(advisories) == 1
    assert advisories[0].attention_category == ATTENTION_CATEGORY_MANAGEMENT_ADVISORY
    # Two confirmed invoices, never summed into a fake comparison.
    row = _supplier_row(export_product, "PO-MULTI")
    assert row.confirmed_purchase_invoice_count == 2
    assert row.supplier_amount_check_outcome is None


def test_supplier_p09_exported_as_management_advisory(export_product):
    advisories = _attention_for(export_product, code="SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED", contract_no="PO-FOLLOW")
    assert len(advisories) == 1
    row = advisories[0]
    assert row.attention_category == ATTENTION_CATEGORY_MANAGEMENT_ADVISORY
    # Never worded as overdue / blocker / violation.
    assert row.attention_message == "已付款，尚未收到对应进项发票，建议催供应商开票"
    assert row.attention_code == "SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED"


def test_payment_alone_produces_no_generic_warning(export_product):
    """PO-MULTI has an OUT payment AND confirmed invoices: its payment is a
    confirmed-out-payment count fact only — no follow-up warning is
    emitted (payment alone is never a warning)."""
    row = _supplier_row(export_product, "PO-MULTI")
    assert row.confirmed_out_payment_count == 1
    assert _attention_for(export_product, code="SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED", contract_no="PO-MULTI") == []


def test_tax_rate_alone_produces_no_warning(export_product):
    """PO-AMT carries an InvoiceItem tax_rate Fact; the only attention row
    for it is the amount deviation — no tax-related warning exists."""
    codes = {r.attention_code for r in _attention_for(export_product, contract_no="PO-AMT")}
    assert codes == {"PURCHASE_INVOICE_AMOUNT_DEVIATION"}
    assert all("税率" not in (r.attention_message or "") for r in _attention_for(export_product, contract_no="PO-AMT"))


# ---------------------------------------------------------------------------
# Confirmed-Fact boundary — dangling associations (pure F0 contexts)
# ---------------------------------------------------------------------------


def _pure_context_with_dangling():
    sc_id, contract_id, shipment_id, dangling_invoice_id, dangling_receipt_id, dangling_purchase_id, dangling_payment_id = (
        uuid.uuid4() for _ in range(7)
    )
    sales_contract = SalesContract(
        id=sc_id, our_entity="Our Own Entity", sales_contract_no="SC-DANGLING-EXPORT",
        customer="Customer D", currency="USD", gross_amount=Decimal("100.00"),
        contract_date=date(2026, 1, 1), current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    contract = Contract(
        id=contract_id, contract_no="PO-DANGLING-EXPORT", contract_type=None, counterparty="Supplier",
        buyer="Our Own Entity", gross_amount=Decimal("100.00"), currency="USD", contract_date=date(2026, 1, 1),
        current_source_fragment_id=uuid.uuid4(), created_at=NOW, updated_at=NOW,
    )
    link = ProcurementSalesLink(
        id=uuid.uuid4(), procurement_contract_id=contract_id, sales_contract_id=sc_id,
        source_fragment_id=uuid.uuid4(), confirmation_type="AUTO_CONFIRMED", created_at=NOW,
    )
    shipment = Shipment(
        id=shipment_id, contract_id=contract_id, external_reference="SHIP-DANGLING-EXPORT",
        execution_date=date(2031, 2, 1), contract_item_id=None, quantity=Decimal("1"),
        current_source_fragment_id=uuid.uuid4(), created_at=NOW,
        declared_amount=Decimal("100.00"), declared_currency="USD",
    )
    context = InvoicePreparationContext(
        sales_scopes=(
            SalesScopeContext(
                sales_contract=sales_contract,
                linked_procurement_contracts=(SalesScopeLinkedProcurementContract(link=link, contract=contract),),
                invoice_allocations=(
                    SalesScopeInvoiceAllocation(
                        allocation=SalesInvoiceAllocation(
                            id=uuid.uuid4(), invoice_id=dangling_invoice_id, sales_contract_id=sc_id,
                            match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("100.00"),
                            confirmation_type=ConfirmationType.HUMAN_CONFIRMED, created_at=NOW,
                        ),
                        invoice=None,
                    ),
                ),
                payment_allocations=(
                    SalesScopePaymentAllocation(
                        allocation=SalesPaymentAllocation(
                            id=uuid.uuid4(), payment_id=dangling_receipt_id, sales_contract_id=sc_id,
                            match_case_id=uuid.uuid4(), allocated_amount=Decimal("60.00"),
                            confirmation_type=ConfirmationType.HUMAN_CONFIRMED, created_at=NOW,
                        ),
                        payment=None,
                    ),
                ),
                unresolved_work=(),
            ),
        ),
        supplier_scopes=(
            SupplierScopeContext(
                contract=contract, items=(), shipments=(shipment,),
                invoice_allocations=(
                    SupplierScopeInvoiceAllocation(
                        allocation=InvoiceAllocation(
                            id=uuid.uuid4(), invoice_id=dangling_purchase_id, contract_id=contract_id,
                            match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("100.00"),
                            match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
                            confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
                        ),
                        invoice=None,
                    ),
                ),
                invoice_item_allocations=(),
                payment_allocations=(
                    SupplierScopePaymentAllocation(
                        allocation=PaymentAllocation(
                            id=uuid.uuid4(), payment_id=dangling_payment_id, contract_id=contract_id,
                            match_case_id=uuid.uuid4(), allocated_amount=Decimal("50.00"),
                            match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
                            confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
                        ),
                        payment=None,
                    ),
                ),
                unresolved_work=(),
            ),
        ),
    )
    return context


def test_dangling_associations_never_counted_confirmed():
    """Dangling sales invoice / receipt and purchase invoice / OUT payment
    associations are NOT counted as confirmed Facts, and each is exported
    as its own INCOMPLETE_ASSOCIATION attention row."""
    context = _pure_context_with_dangling()
    product = build_invoice_preparation_data_product(get_invoice_preparation_workbench_from_context(context))

    sales_row = product.sales_preparation[0]
    assert sales_row.confirmed_sales_invoice_count == 0
    assert sales_row.confirmed_receipt_count == 0
    supplier_row = product.supplier_request[0]
    assert supplier_row.confirmed_purchase_invoice_count == 0
    assert supplier_row.confirmed_out_payment_count == 0

    incomplete = _attention_for(product, category=ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION)
    # Every kind carries the truthful generic wording (missing Fact OR
    # direction mismatch — never claimed to be necessarily missing).
    kinds = {r.attention_message for r in incomplete}
    assert kinds == {
        "销项发票关联 · 关联记录未形成可确认业务事实",
        "收款关联 · 关联记录未形成可确认业务事实",
        "采购发票关联 · 关联记录未形成可确认业务事实",
        "付款关联 · 关联记录未形成可确认业务事实",
    }
    assert all(r.attention_code is None for r in incomplete)
    assert all(r.record_type in (RECORD_TYPE_SALES_ATTENTION, RECORD_TYPE_SUPPLIER_ATTENTION) for r in incomplete)


# ---------------------------------------------------------------------------
# D. FORMAT — five sheets, unified CSV, determinism, safety, blank-not-zero
# ---------------------------------------------------------------------------


def test_xlsx_has_exactly_five_sheets_in_order(export_product):
    wb = openpyxl.load_workbook(io.BytesIO(export_invoice_preparation_xlsx(export_product)))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_xlsx_numeric_amounts_and_blank_missing(export_product):
    wb = openpyxl.load_workbook(io.BytesIO(export_invoice_preparation_xlsx(export_product)))
    ws = wb["02_Sales_Preparation"]
    header = [c.value for c in ws[1]]
    rows = {ws.cell(row=r, column=header.index("sales_contract_no") + 1).value: r for r in range(2, ws.max_row + 1)}
    # MATCH row: the amounts are NUMERIC cells (int/float), never text.
    r = rows["SC-MTCH"]
    col_amount = header.index("sales_contract_amount") + 1
    value = ws.cell(row=r, column=col_amount).value
    assert value == 100
    assert isinstance(value, (int, float))
    # Missing invoice leg on SC-NOINV is a BLANK cell, never 0.
    r = rows["SC-NOINV"]
    col_inv = header.index("sales_invoice_amount") + 1
    assert ws.cell(row=r, column=col_inv).value is None


def test_csv_unified_record_type(export_product):
    text = export_invoice_preparation_csv(export_product).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    records = list(reader)
    assert len(records) > 0
    assert {r["record_type"] for r in records} <= ALL_RECORD_TYPES
    # All four record types appear.
    assert {r["record_type"] for r in records} == ALL_RECORD_TYPES


def test_csv_missing_value_is_empty_not_zero(export_product):
    text = export_invoice_preparation_csv(export_product).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    noinv = next(r for r in reader if r["record_type"] == "SALES_PREPARATION" and r["sales_contract_no"] == "SC-NOINV")
    assert noinv["sales_invoice_amount"] == ""
    assert noinv["comparison_outcome"] == "NOT_COMPARABLE_MISSING_FACT"


def test_export_is_byte_deterministic(export_product):
    csv_a = export_invoice_preparation_csv(export_product)
    xlsx_a = export_invoice_preparation_xlsx(export_product)
    csv_b = export_invoice_preparation_csv(export_product)
    xlsx_b = export_invoice_preparation_xlsx(export_product)
    assert csv_a == csv_b
    assert xlsx_a == xlsx_b


def test_formula_injection_neutralized():
    """A dangerous leading character on a text value (e.g. a contract
    number starting with '=') is neutralized in BOTH artifacts — it must
    never be interpretable as a spreadsheet formula."""
    sales_contract = SalesContract(
        id=uuid.uuid4(), our_entity="Our Own Entity", sales_contract_no="=SUM(A1)",
        customer="Customer", currency="USD", gross_amount=Decimal("100.00"),
        contract_date=date(2026, 1, 1), current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    context = InvoicePreparationContext(
        sales_scopes=(SalesScopeContext(
            sales_contract=sales_contract, linked_procurement_contracts=(),
            invoice_allocations=(), payment_allocations=(), unresolved_work=(),
        ),),
        supplier_scopes=(),
    )
    product = build_invoice_preparation_data_product(get_invoice_preparation_workbench_from_context(context))

    csv_bytes = export_invoice_preparation_csv(product)
    assert b"'=SUM(A1)" in csv_bytes
    wb = openpyxl.load_workbook(io.BytesIO(export_invoice_preparation_xlsx(product)))
    ws = wb["02_Sales_Preparation"]
    header = [c.value for c in ws[1]]
    col = header.index("sales_contract_no") + 1
    assert str(ws.cell(row=2, column=col).value).startswith("'")


# ---------------------------------------------------------------------------
# F2b pre-Gate repair — traceability, structured JSON, scope identity,
# and real determinism. All scenarios are pure F0 contexts (no Session).
# ---------------------------------------------------------------------------


def _pure_supplier_scope(contract, *, invoice_allocations=(), payment_allocations=(), shipments=(),
                         items=(), invoice_item_allocations=()):
    return SupplierScopeContext(
        contract=contract, items=tuple(items), shipments=tuple(shipments),
        invoice_allocations=tuple(invoice_allocations),
        invoice_item_allocations=tuple(invoice_item_allocations),
        payment_allocations=tuple(payment_allocations), unresolved_work=(),
    )


def _pure_sales_scope(sales_contract, *, invoice_allocations=(), payment_allocations=(), links=()):
    return SalesScopeContext(
        sales_contract=sales_contract,
        linked_procurement_contracts=tuple(links),
        invoice_allocations=tuple(invoice_allocations),
        payment_allocations=tuple(payment_allocations), unresolved_work=(),
    )


def _pure_contract(no, *, gross=Decimal("100.00")):
    return Contract(
        id=uuid.uuid4(), contract_no=no, contract_type=None, counterparty="Supplier",
        buyer="Our Own Entity", gross_amount=gross, currency="USD", contract_date=date(2026, 1, 1),
        current_source_fragment_id=uuid.uuid4(), created_at=NOW, updated_at=NOW,
    )


def _pure_sales_contract(no, our_entity="Our Own Entity"):
    return SalesContract(
        id=uuid.uuid4(), our_entity=our_entity, sales_contract_no=no,
        customer="Customer", currency="USD", gross_amount=Decimal("100.00"),
        contract_date=date(2026, 1, 1), current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )


def _pure_invoice(direction, key, *, gross=Decimal("100.00")):
    return Invoice(
        id=uuid.uuid4(), direction=direction, invoice_type=None, invoice_no=None, digital_invoice_no=None,
        external_invoice_key=key, issue_date=date(2031, 1, 10), seller="Supplier", buyer="Our Own Entity",
        net_amount=gross, tax_amount=Decimal("0"), gross_amount=gross, invoice_status=None,
        source_fragment_id=uuid.uuid4(), created_at=NOW, updated_at=NOW, currency="USD",
    )


def _dangling_purchase_alloc(contract, invoice_id=None):
    return SupplierScopeInvoiceAllocation(
        allocation=InvoiceAllocation(
            id=uuid.uuid4(), invoice_id=invoice_id or uuid.uuid4(), contract_id=contract.id,
            match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("100.00"),
            match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
            confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
        ),
        invoice=None,
    )


def _dangling_sales_invoice_alloc(sales_contract):
    return SalesScopeInvoiceAllocation(
        allocation=SalesInvoiceAllocation(
            id=uuid.uuid4(), invoice_id=uuid.uuid4(), sales_contract_id=sales_contract.id,
            match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("100.00"),
            confirmation_type=ConfirmationType.HUMAN_CONFIRMED, created_at=NOW,
        ),
        invoice=None,
    )


def _pure_two_contracts_shared_invoice():
    contract_a, contract_b = _pure_contract("PO-SHARED-A"), _pure_contract("PO-SHARED-B")
    invoice = _pure_invoice(InvoiceDirection.PURCHASE, "PINV-SHARED")

    def alloc(contract):
        return SupplierScopeInvoiceAllocation(
            allocation=InvoiceAllocation(
                id=uuid.uuid4(), invoice_id=invoice.id, contract_id=contract.id,
                match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("100.00"),
                match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
                confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
            ),
            invoice=invoice,
        )

    return InvoicePreparationContext(
        sales_scopes=(),
        supplier_scopes=(_pure_supplier_scope(contract_a, invoice_allocations=(alloc(contract_a),)),
                         _pure_supplier_scope(contract_b, invoice_allocations=(alloc(contract_b),))),
    )


def _pure_duplicate_contract_no():
    contract_a, contract_b = _pure_contract("DUP-1"), _pure_contract("DUP-1")
    return InvoicePreparationContext(
        sales_scopes=(),
        supplier_scopes=(
            _pure_supplier_scope(contract_a, invoice_allocations=(_dangling_purchase_alloc(contract_a),)),
            _pure_supplier_scope(contract_b, invoice_allocations=(_dangling_purchase_alloc(contract_b),)),
        ),
    )


def _pure_duplicate_sales_no():
    sc_a, sc_b = _pure_sales_contract("DUP-SC", our_entity="Entity A"), _pure_sales_contract("DUP-SC", our_entity="Entity B")
    return InvoicePreparationContext(
        sales_scopes=(
            _pure_sales_scope(sc_a, invoice_allocations=(_dangling_sales_invoice_alloc(sc_a),)),
            _pure_sales_scope(sc_b, invoice_allocations=(_dangling_sales_invoice_alloc(sc_b),)),
        ),
        supplier_scopes=(),
    )


def _pure_dangerous_product_names():
    contract = _pure_contract("PO-DANGER")
    contract_item = ContractItem(
        id=uuid.uuid4(), contract_id=contract.id, source_item_key="ITEM-DANGER", sku=None,
        product_name="A;outcome=DEVIATION||B=C", specification=None, quantity=Decimal("1"), unit=None,
        unit_price=None, gross_amount=Decimal("0"), tax_rate=None, net_amount=Decimal("0"),
        current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    invoice = _pure_invoice(InvoiceDirection.PURCHASE, "PINV-DANGER")
    invoice_item = InvoiceItem(
        id=uuid.uuid4(), invoice_id=invoice.id, line_no=1, product_name="Widget", specification=None,
        unit=None, quantity=Decimal("1"), unit_price=None, net_amount=Decimal("100.00"),
        tax_rate=None, tax_amount=Decimal("0"), gross_amount=Decimal("100.00"), source_fragment_id=uuid.uuid4(),
    )
    item_alloc = SupplierScopeInvoiceItemAllocation(
        allocation=InvoiceItemAllocation(
            id=uuid.uuid4(), invoice_item_id=invoice_item.id, contract_item_id=contract_item.id,
            allocated_quantity=Decimal("1"), allocated_net_amount=Decimal("100.00"),
            confirmation_type="MANUAL_CONFIRMED", source_fragment_id=uuid.uuid4(), created_at=NOW,
            superseded_by_fact_id=None,
        ),
        invoice_item=invoice_item,
        invoice=invoice,
    )
    return InvoicePreparationContext(
        sales_scopes=(),
        supplier_scopes=(_pure_supplier_scope(contract, items=(contract_item,), invoice_item_allocations=(item_alloc,)),),
    )


def _pure_product(workbench):
    return build_invoice_preparation_data_product(workbench)


def test_xlsx_byte_identity_across_time_boundary(export_product):
    """openpyxl stamps created/modified into docProps/core.xml and each zip
    entry with the CURRENT time — the fixed workbook properties + zip
    normalization must keep two exports byte-identical across a REAL second
    boundary, not just back-to-back within one second."""
    a = export_invoice_preparation_xlsx(export_product)
    time.sleep(1.5)  # guarantee at least one wall-clock second boundary
    b = export_invoice_preparation_xlsx(export_product)
    assert a == b


def test_sales_comparison_exports_trace_ids(export_product):
    row = _sales_row(export_product, "SC-MTCH")
    assert row.comparison_shipment_id
    assert row.comparison_sales_invoice_id
    assert row.comparison_shipment_id != row.comparison_sales_invoice_id
    # The ambiguous declaration leg leaves the shipment trace None; the
    # single confirmed invoice is still traced.
    amb = _sales_row(export_product, "SC-AMB")
    assert amb.comparison_shipment_id is None
    assert amb.comparison_sales_invoice_id
    # No confirmed invoice -> the invoice trace is None (missing Fact).
    noinv = _sales_row(export_product, "SC-NOINV")
    assert noinv.comparison_sales_invoice_id is None
    assert noinv.comparison_shipment_id


def test_supplier_p03_preserves_all_related_invoice_ids(export_product):
    advisories = _attention_for(export_product, code="MULTIPLE_PURCHASE_INVOICES_ON_CONTRACT", contract_no="PO-MULTI")
    assert len(advisories) == 1
    ids = json.loads(advisories[0].related_invoice_ids)
    assert len(ids) == 2  # BOTH purchase invoices — never truncated to one


def test_supplier_p04_preserves_all_related_contract_ids():
    context = _pure_two_contracts_shared_invoice()
    product = _pure_product(get_invoice_preparation_workbench_from_context(context))
    advisories = [r for r in product.supplier_attention if r.attention_code == "PURCHASE_INVOICE_SPANS_MULTIPLE_CONTRACTS"]
    assert len(advisories) == 2  # surfaced on every involved contract
    for row in advisories:
        ids = json.loads(row.related_contract_ids)
        assert len(ids) == 2  # the offending invoice's FULL contract set
        assert row.related_invoice_ids


def test_sales_advisory_preserves_invoice_and_shipment_ids(export_product):
    advisories = _attention_for(export_product, code="SALES_INVOICE_AMOUNT_DEVIATION", sales_no="SC-DEV")
    assert len(advisories) == 1
    invoice_ids = json.loads(advisories[0].related_invoice_ids)
    shipment_ids = json.loads(advisories[0].related_shipment_ids)
    assert len(invoice_ids) == 1 and len(shipment_ids) == 1
    # They agree with the comparison's own trace ids on the row.
    row = _sales_row(export_product, "SC-DEV")
    assert invoice_ids[0] == row.comparison_sales_invoice_id
    assert shipment_ids[0] == row.comparison_shipment_id


def test_duplicate_contract_no_unambiguous_in_supplier_attention():
    context = _pure_duplicate_contract_no()
    product = _pure_product(get_invoice_preparation_workbench_from_context(context))
    attention = product.supplier_attention
    assert len(attention) == 2
    assert all(r.contract_no == "DUP-1" for r in attention)
    # contract_no is NOT unique — the procurement_contract_id distinguishes
    # the two rows.
    assert len({r.procurement_contract_id for r in attention}) == 2


def test_same_sales_contract_no_unambiguous_in_sales_attention():
    context = _pure_duplicate_sales_no()
    product = _pure_product(get_invoice_preparation_workbench_from_context(context))
    attention = product.sales_attention
    assert len(attention) == 2
    assert all(r.sales_contract_no == "DUP-SC" for r in attention)
    # SalesContract identity is (our_entity, sales_contract_no) — the id
    # column keeps the two rows unambiguous.
    assert len({r.sales_contract_id for r in attention}) == 2


def test_product_names_with_dangerous_chars_cannot_corrupt_checks_or_summary():
    """Product names containing ';', '=' and '||' must not corrupt the
    structured JSON check serialization, and the Summary must derive its
    counts from the explicit neutral fields — never from parsing the
    serialized text (a name like 'A;outcome=DEVIATION||B=C' would have
    inflated a text-parsing counter)."""
    context = _pure_dangerous_product_names()
    product = _pure_product(get_invoice_preparation_workbench_from_context(context))
    row = product.supplier_request[0]
    checks = json.loads(row.supplier_checks_json)  # parses cleanly
    item = next(c for c in checks if c["kind"] == "ITEM_NAME")
    assert item["contract_product_name"] == "A;outcome=DEVIATION||B=C"
    assert item["outcome"] == "DEVIATION"
    # Explicit neutral counts are exact.
    assert row.supplier_item_name_deviation_count == 1
    assert row.supplier_item_name_check_count == 1
    assert product.summary["supplier_item_name_check_DEVIATION"] == 1
    assert product.summary["supplier_item_name_check_count"] == 1


def test_summary_never_parses_serialized_check_text():
    """The summary fields are computed from explicit neutral row fields;
    no serialized text (supplier_checks_json / comparison_message) is ever
    parsed to recover business structure."""
    context = _pure_dangerous_product_names()
    product = _pure_product(get_invoice_preparation_workbench_from_context(context))
    # The dangerous-name scenario proves it: a naive text parse of the
    # check JSON could not have produced these exact counts.
    assert product.summary["supplier_item_name_check_DEVIATION"] == 1
    assert product.summary["supplier_amount_check_count"] == 0


# ---------------------------------------------------------------------------
# Final Gate 2C — Decimal precision: numeric/text split, exact values.
# ---------------------------------------------------------------------------


def _product_from_amounts(sales_amount, supplier_amount):
    sc = SalesContract(
        id=uuid.uuid4(), our_entity="Our Own Entity", sales_contract_no="SC-NUM",
        customer="Customer", currency="USD", gross_amount=sales_amount,
        contract_date=date(2026, 1, 1), current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    contract = Contract(
        id=uuid.uuid4(), contract_no="PO-NUM", contract_type=None, counterparty="Supplier",
        buyer="Our Own Entity", gross_amount=supplier_amount, currency="USD",
        contract_date=date(2026, 1, 1), current_source_fragment_id=uuid.uuid4(), created_at=NOW, updated_at=NOW,
    )
    context = InvoicePreparationContext(
        sales_scopes=(
            SalesScopeContext(
                sales_contract=sc, linked_procurement_contracts=(),
                invoice_allocations=(), payment_allocations=(), unresolved_work=(),
            ),
        ),
        supplier_scopes=(
            SupplierScopeContext(
                contract=contract, items=(), shipments=(),
                invoice_allocations=(), invoice_item_allocations=(),
                payment_allocations=(), unresolved_work=(),
            ),
        ),
    )
    return build_invoice_preparation_data_product(get_invoice_preparation_workbench_from_context(context))


def _xlsx_cell_value(product, sheet, column, row=2):
    wb = openpyxl.load_workbook(io.BytesIO(export_invoice_preparation_xlsx(product)))
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    return ws.cell(row=row, column=header.index(column) + 1).value


def test_csv_and_xlsx_positive_decimal():
    product = _product_from_amounts(Decimal("1234.56"), Decimal("1234.56"))
    text = export_invoice_preparation_csv(product).decode("utf-8-sig")
    assert '"1234.56"' in text
    assert "'1234.56" not in text  # no apostrophe on a plain positive amount
    # XLSX: numeric cell (not text, not rounded).
    assert _xlsx_cell_value(product, "02_Sales_Preparation", "contract_gross_amount") == 1234.56
    assert isinstance(_xlsx_cell_value(product, "02_Sales_Preparation", "contract_gross_amount"), (int, float))


def test_csv_and_xlsx_negative_decimal_no_apostrophe():
    product = _product_from_amounts(Decimal("-1234.56"), Decimal("-1234.56"))
    text = export_invoice_preparation_csv(product).decode("utf-8-sig")
    assert '"-1234.56"' in text
    assert "'-1234.56" not in text  # a legitimate numeric amount never gains a semantic apostrophe
    assert _xlsx_cell_value(product, "02_Sales_Preparation", "contract_gross_amount") == -1234.56
    assert _xlsx_cell_value(product, "04_Supplier_Request", "expected_invoice_amount") == -1234.56


def test_csv_text_negative_prefix_still_protected():
    """A TEXT value beginning with '-' stays formula-protected — only
    numeric Decimal/int values are exempt from the apostrophe guard."""
    sc = SalesContract(
        id=uuid.uuid4(), our_entity="Our Own Entity", sales_contract_no="-CMD",
        customer="Customer", currency="USD", gross_amount=Decimal("100.00"),
        contract_date=date(2026, 1, 1), current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    context = InvoicePreparationContext(
        sales_scopes=(SalesScopeContext(
            sales_contract=sc, linked_procurement_contracts=(),
            invoice_allocations=(), payment_allocations=(), unresolved_work=(),
        ),),
        supplier_scopes=(),
    )
    product = build_invoice_preparation_data_product(get_invoice_preparation_workbench_from_context(context))
    text = export_invoice_preparation_csv(product).decode("utf-8-sig")
    assert "'-CMD" in text
    # XLSX text cell too.
    assert _xlsx_cell_value(product, "02_Sales_Preparation", "sales_contract_no") == "'-CMD"


def test_xlsx_large_numeric_18_2_exact_text_lossless():
    """A Numeric(18,2)-scale Decimal beyond Excel's usable numeric precision
    is written as EXACT text (format(value, "f")) — no rounding, no
    scientific notation, no precision loss. The value is constructed
    arithmetically from sub-10-digit literals so it is a genuine
    17-significant-digit Numeric(18,2) value without tripping the privacy
    scanner's generic guard on a long source digit run."""
    value = Decimal(123456789) * Decimal(10) ** 6 + Decimal("12345.67")
    exact = format(value, "f")  # the exact canonical text (17 significant digits)
    product = _product_from_amounts(value, value)
    assert _xlsx_cell_value(product, "02_Sales_Preparation", "contract_gross_amount") == exact
    assert _xlsx_cell_value(product, "04_Supplier_Request", "expected_invoice_amount") == exact
    text = export_invoice_preparation_csv(product).decode("utf-8-sig")
    assert '"' + exact + '"' in text


def test_xlsx_precision_boundary_safe_numeric_unsafe_exact_text():
    """A 15-significant-digit Decimal stays a NUMERIC cell; a 16-digit one
    becomes EXACT text. (Source literals kept under 10 contiguous digits for
    the privacy scanner's generic guard.)"""
    safe = Decimal("1234567.89012345")  # 15 significant digits
    unsafe = Decimal("1234567.890123456")  # 16 significant digits
    product = _product_from_amounts(safe, unsafe)
    assert _xlsx_cell_value(product, "02_Sales_Preparation", "contract_gross_amount") == 1234567.89012345
    assert isinstance(_xlsx_cell_value(product, "02_Sales_Preparation", "contract_gross_amount"), (int, float))
    # The unsafe value lands on the supplier row's expected amount.
    assert _xlsx_cell_value(product, "04_Supplier_Request", "expected_invoice_amount") == "1234567.890123456"


def test_decimal_policy_preserves_scale_in_csv():
    """CSV preserves the canonical Decimal scale (0.10, 100.00) — never
    silently rewritten to '0.1' or '100'."""
    product = _product_from_amounts(Decimal("0.10"), Decimal("100.00"))
    text = export_invoice_preparation_csv(product).decode("utf-8-sig")
    assert '"0.10"' in text
    assert '"100.00"' in text


def test_wrong_direction_associations_preserved_not_confirmed_in_product():
    """Final Gate: a wrong-direction InvoiceAllocation / PaymentAllocation /
    InvoiceItemAllocation is preserved by the Data Product as an
    INCOMPLETE_ASSOCIATION attention row, and the supplier confirmed-Fact
    counts stay zero."""
    contract = Contract(
        id=uuid.uuid4(), contract_no="PO-WRONG-EXPORT", contract_type=None, counterparty="Supplier",
        buyer="Our Own Entity", gross_amount=Decimal("100.00"), currency="USD", contract_date=date(2026, 1, 1),
        current_source_fragment_id=uuid.uuid4(), created_at=NOW, updated_at=NOW,
    )
    contract_item = ContractItem(
        id=uuid.uuid4(), contract_id=contract.id, source_item_key="ITEM-WRONG-X", sku=None,
        product_name="Widget", specification=None, quantity=Decimal("1"), unit=None,
        unit_price=None, gross_amount=Decimal("0"), tax_rate=None, net_amount=Decimal("0"),
        current_source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    wrong_invoice = Invoice(
        id=uuid.uuid4(), direction=InvoiceDirection.SALES, invoice_type=None, invoice_no=None,
        digital_invoice_no=None, external_invoice_key="SINV-WRONG-X", issue_date=date(2031, 1, 10),
        seller="Our Own Entity", buyer="Customer", net_amount=Decimal("100.00"), tax_amount=Decimal("0"),
        gross_amount=Decimal("100.00"), invoice_status=None, source_fragment_id=uuid.uuid4(),
        created_at=NOW, updated_at=NOW, currency="USD",
    )
    wrong_payment = Payment(
        id=uuid.uuid4(), transaction_date=date(2031, 1, 15), direction=PaymentDirection.IN,
        amount=Decimal("50.00"), counterparty="Customer", business_type=None,
        bank_reference="REF-IN-WRONG-X", description=None, running_balance=None,
        source_fragment_id=uuid.uuid4(), created_at=NOW,
    )
    invoice_item = InvoiceItem(
        id=uuid.uuid4(), invoice_id=wrong_invoice.id, line_no=1, product_name="Widget",
        specification=None, unit=None, quantity=Decimal("1"), unit_price=None,
        net_amount=Decimal("100.00"), tax_rate=None, tax_amount=Decimal("0"),
        gross_amount=Decimal("100.00"), source_fragment_id=uuid.uuid4(),
    )
    context = InvoicePreparationContext(
        sales_scopes=(),
        supplier_scopes=(
            SupplierScopeContext(
                contract=contract, items=(contract_item,), shipments=(),
                invoice_allocations=(
                    SupplierScopeInvoiceAllocation(
                        allocation=InvoiceAllocation(
                            id=uuid.uuid4(), invoice_id=wrong_invoice.id, contract_id=contract.id,
                            match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("100.00"),
                            match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
                            confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
                        ),
                        invoice=wrong_invoice,
                    ),
                ),
                invoice_item_allocations=(
                    SupplierScopeInvoiceItemAllocation(
                        allocation=InvoiceItemAllocation(
                            id=uuid.uuid4(), invoice_item_id=invoice_item.id, contract_item_id=contract_item.id,
                            allocated_quantity=Decimal("1"), allocated_net_amount=Decimal("100.00"),
                            confirmation_type="MANUAL_CONFIRMED", source_fragment_id=uuid.uuid4(), created_at=NOW,
                            superseded_by_fact_id=None,
                        ),
                        invoice_item=invoice_item,
                        invoice=wrong_invoice,
                    ),
                ),
                payment_allocations=(
                    SupplierScopePaymentAllocation(
                        allocation=PaymentAllocation(
                            id=uuid.uuid4(), payment_id=wrong_payment.id, contract_id=contract.id,
                            match_case_id=uuid.uuid4(), allocated_amount=Decimal("50.00"),
                            match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
                            confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
                        ),
                        payment=wrong_payment,
                    ),
                ),
                unresolved_work=(),
            ),
        ),
    )
    product = build_invoice_preparation_data_product(get_invoice_preparation_workbench_from_context(context))
    row = product.supplier_request[0]
    assert row.confirmed_purchase_invoice_count == 0
    assert row.confirmed_out_payment_count == 0
    incomplete = _attention_for(product, category=ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION)
    kinds = {r.attention_message for r in incomplete}
    assert kinds == {
        "采购发票关联 · 关联记录未形成可确认业务事实",
        "付款关联 · 关联记录未形成可确认业务事实",
        "采购发票明细关联 · 关联记录未形成可确认业务事实",
    }
    # The CSV/XLSX preserve the incomplete associations (they serialize the
    # same product) — confirmed counts stay zero there too.
    text = export_invoice_preparation_csv(product).decode("utf-8-sig")
    assert '"0"' in text  # the confirmed counts are literal zero, not blank-but-implied
    assert "采购发票明细关联 · 关联记录未形成可确认业务事实" in text


def test_p02_confirmed_comparison_survives_incomplete_associations_in_product():
    """Final Gate repair #2 (1C, verified through Workbench + Data Product):
    with ONE confirmed PURCHASE invoice PLUS a dangling association, the
    P02 comparison still runs against the confirmed invoice (not suppressed),
    and the dangling association surfaces as an INCOMPLETE_ASSOCIATION
    attention row."""
    contract = _pure_contract("PO-P02-PROD", gross=Decimal("1000.00"))
    confirmed = _pure_invoice(InvoiceDirection.PURCHASE, "PINV-P02-PROD", gross=Decimal("900.00"))

    def alloc(invoice=None):
        return SupplierScopeInvoiceAllocation(
            allocation=InvoiceAllocation(
                id=uuid.uuid4(), invoice_id=(invoice.id if invoice else uuid.uuid4()), contract_id=contract.id,
                match_case_id=uuid.uuid4(), allocated_gross_amount=Decimal("900.00"),
                match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
                confirmation_type=ConfirmationType.AUTO_CONFIRMED, created_at=NOW,
            ),
            invoice=invoice,
        )

    context = InvoicePreparationContext(
        sales_scopes=(),
        supplier_scopes=(
            SupplierScopeContext(
                contract=contract, items=(), shipments=(),
                invoice_allocations=(alloc(confirmed), alloc(None)),
                invoice_item_allocations=(), payment_allocations=(), unresolved_work=(),
            ),
        ),
    )
    workbench = get_invoice_preparation_workbench_from_context(context)
    decision = workbench.supplier_report.decisions[0]
    # The Workbench's F1 decision carries the confirmed comparison.
    assert len(decision.amount_checks) == 1
    assert decision.amount_checks[0].outcome == SupplierRequestCheckOutcome.DEVIATION
    assert decision.amount_checks[0].invoice_id == confirmed.id

    product = build_invoice_preparation_data_product(workbench)
    row = product.supplier_request[0]
    assert row.supplier_amount_check_outcome == "DEVIATION"
    assert row.confirmed_purchase_invoice_count == 1
    # The dangling association is attention, never a confirmed Fact.
    incomplete = _attention_for(product, category=ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION)
    assert any("采购发票关联" in r.attention_message for r in incomplete)

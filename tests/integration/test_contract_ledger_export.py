"""Phase 2D.1-R4 — Contract Business Ledger CSV/XLSX Data Product.

Covers export parity with the Application-layer projection (section 39),
CSV/XLSX structural validity (sections 40-41), and the HARD
formula-injection security requirement (section 30).
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from bel.application.contract_business_ledger import ContractLedgerFilters, get_contract_business_ledger
from bel.application.contract_ledger_export import (
    export_contract_business_ledger_csv,
    export_contract_business_ledger_xlsx,
)
from bel.application.sales_contract_facts import create_sales_contract_fact
from bel.infrastructure.persistence.database import make_engine, make_session_factory
from bel.infrastructure.persistence.models import Base
from bel.infrastructure.persistence.repositories import ContractRepository, EvidenceRepository
from bel.domain.contract import Contract
from bel.domain.evidence import EvidenceDocument, EvidenceFragment, FragmentKind
from bel.application.procurement_sales_link import add_procurement_sales_link
from bel.domain.procurement_sales_link import ConfirmationType as LinkConfirmationType

NOW = datetime.now(timezone.utc)


@pytest.fixture
def db_session():
    engine = make_engine(":memory:")
    Base.metadata.create_all(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as session:
        yield session


def _make_fragment(session, raw_data=None):
    doc = EvidenceDocument(
        id=uuid.uuid4(), file_name="x", sha256=uuid.uuid4().hex + uuid.uuid4().hex, source_type="t", imported_at=NOW
    )
    evidence_repo = EvidenceRepository(session)
    evidence_repo.add_document(doc)
    frag = EvidenceFragment(
        id=uuid.uuid4(),
        evidence_document_id=doc.id,
        fragment_kind=FragmentKind.MANUAL_FACT,
        sheet_name=None,
        row_number=None,
        locator_json={"section": "test", "index": 0},
        raw_data=raw_data or {},
        created_at=NOW,
    )
    evidence_repo.add_fragment(frag)
    session.flush()
    return frag


def _make_contract(session, fragment_id, contract_no=None, counterparty="Supplier A", buyer="Our Own Entity"):
    contract = Contract(
        id=uuid.uuid4(),
        contract_no=contract_no or f"C-{uuid.uuid4().hex[:8]}",
        contract_type=None,
        counterparty=counterparty,
        buyer=buyer,
        gross_amount=Decimal("1000.00"),
        currency="CNY",
        contract_date=date(2026, 1, 1),
        current_source_fragment_id=fragment_id,
        created_at=NOW,
        updated_at=NOW,
    )
    ContractRepository(session).add(contract)
    session.flush()
    return contract


def _make_sales_contract(session, fragment_id, sales_contract_no=None, our_entity="Our Own Entity", customer=None):
    result = create_sales_contract_fact(
        session,
        our_entity=our_entity,
        sales_contract_no=sales_contract_no or f"SC-{uuid.uuid4().hex[:8]}",
        fields={"customer": customer} if customer else {},
        source_fragment_id=fragment_id,
        created_at=NOW,
    )
    return result.sales_contract


def _link(session, contract, sales_contract):
    frag = _make_fragment(session)
    result = add_procurement_sales_link(
        session,
        procurement_contract_id=contract.id,
        sales_contract_id=sales_contract.id,
        source_fragment_id=frag.id,
        confirmation_type=LinkConfirmationType.AUTO_CONFIRMED,
        created_at=NOW,
    )
    return result.link


# ---------------------------------------------------------------------------
# Y/Z — export parity with the Application projection
# ---------------------------------------------------------------------------


def test_csv_parity_with_application_rows(db_session):
    frag = _make_fragment(db_session)
    c1 = _make_contract(db_session, frag.id, contract_no="PO-0001")
    c2 = _make_contract(db_session, frag.id, contract_no="PO-0002")
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    csv_bytes = export_contract_business_ledger_csv(ledger)
    text = csv_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    csv_ids = [row["contract_id"] for row in reader]
    app_ids = [str(r.contract.id) for r in ledger.rows]
    assert csv_ids == app_ids


def test_xlsx_parity_with_application_rows(db_session):
    frag = _make_fragment(db_session)
    _make_contract(db_session, frag.id, contract_no="PO-0001")
    _make_contract(db_session, frag.id, contract_no="PO-0002")
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    xlsx_bytes = export_contract_business_ledger_xlsx(ledger)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Contract Business Ledger"]
    headers = [c.value for c in ws[1]]
    contract_id_col = headers.index("contract_id")
    xlsx_ids = [row[contract_id_col].value for row in ws.iter_rows(min_row=2)]
    app_ids = [str(r.contract.id) for r in ledger.rows]
    assert xlsx_ids == app_ids


def test_export_filters_match_page_filters(db_session):
    frag = _make_fragment(db_session)
    c1 = _make_contract(db_session, frag.id, contract_no="PO-0001", counterparty="Alpha Supplier")
    c2 = _make_contract(db_session, frag.id, contract_no="PO-0002", counterparty="Beta Supplier")
    db_session.commit()

    filters = ContractLedgerFilters(supplier="alpha")
    ledger = get_contract_business_ledger(db_session, filters)
    assert {r.contract.id for r in ledger.rows} == {c1.id}

    csv_bytes = export_contract_business_ledger_csv(ledger)
    reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig")))
    csv_ids = {row["contract_id"] for row in reader}
    assert csv_ids == {str(c1.id)}

    xlsx_bytes = export_contract_business_ledger_xlsx(ledger)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Contract Business Ledger"]
    headers = [c.value for c in ws[1]]
    contract_id_col = headers.index("contract_id")
    xlsx_ids = {row[contract_id_col].value for row in ws.iter_rows(min_row=2)}
    assert xlsx_ids == {str(c1.id)}


# ---------------------------------------------------------------------------
# AA/AB — formula injection (HARD security)
# ---------------------------------------------------------------------------


DANGEROUS_VALUES = ["=1+1", "+SUM(A1:A2)", "-1+2", "@cmd|'/c calc'!A1", "\tformula", "\rformula"]


@pytest.mark.parametrize("dangerous", DANGEROUS_VALUES)
def test_csv_neutralizes_formula_injection(db_session, dangerous):
    frag = _make_fragment(db_session)
    contract = _make_contract(db_session, frag.id, contract_no="PO-INJ", counterparty=dangerous)
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    csv_bytes = export_contract_business_ledger_csv(ledger)
    reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig")))
    rows = list(reader)
    row = next(r for r in rows if r["contract_id"] == str(contract.id))
    cell = row["supplier_counterparty"]
    assert not cell.startswith(("=", "+", "-", "@"))
    # The original text content is preserved (minus the neutralizing
    # marker), never silently dropped.
    assert dangerous.lstrip("=+-@\t\r") in cell


@pytest.mark.parametrize("dangerous", DANGEROUS_VALUES)
def test_xlsx_neutralizes_formula_injection(db_session, dangerous):
    frag = _make_fragment(db_session)
    contract = _make_contract(db_session, frag.id, contract_no="PO-INJ2", counterparty=dangerous)
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    xlsx_bytes = export_contract_business_ledger_xlsx(ledger)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Contract Business Ledger"]
    headers = [c.value for c in ws[1]]
    col = headers.index("supplier_counterparty")
    values = [row[col].value for row in ws.iter_rows(min_row=2)]
    target = next(v for v in values if v is not None and dangerous.lstrip("=+-@\t\r") in v)
    assert not target.startswith(("=", "+", "-", "@"))
    # openpyxl must have written a literal string cell, never a formula.
    for row in ws.iter_rows(min_row=2):
        assert row[col].data_type != "f"


def test_json_cell_is_also_neutralized_and_parseable(db_session):
    frag = _make_fragment(db_session)
    contract = _make_contract(db_session, frag.id, contract_no="PO-ITEMJSON")
    from bel.application.contract_item_facts import create_contract_item_fact

    create_contract_item_fact(
        db_session,
        contract_id=contract.id,
        source_item_key="ITEM-A",
        fields={"product_name": "=cmd|' /c calc'!A0", "quantity": Decimal("10")},
        source_fragment_id=frag.id,
        created_at=NOW,
    )
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    csv_bytes = export_contract_business_ledger_csv(ledger)
    reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig")))
    row = next(r for r in reader if r["contract_id"] == str(contract.id))
    raw_cell = row["items_json"]
    assert not raw_cell.startswith(("=", "+", "-", "@"))
    parsed = json.loads(raw_cell.lstrip("'"))
    assert parsed[0]["product_name"] == "=cmd|' /c calc'!A0"


# ---------------------------------------------------------------------------
# No-cross-bridge aggregation, reflected in the export (section 34/28)
# ---------------------------------------------------------------------------


def test_sales_scope_export_includes_contract_date(db_session):
    frag = _make_fragment(db_session)
    contract = _make_contract(db_session, frag.id, contract_no="PO-DATE")
    result = create_sales_contract_fact(
        db_session,
        our_entity="Our Own Entity",
        sales_contract_no="SC-DATE",
        fields={"customer": "Customer", "contract_date": date(2026, 4, 1)},
        source_fragment_id=frag.id,
        created_at=NOW,
    )
    _link(db_session, contract, result.sales_contract)
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    csv_bytes = export_contract_business_ledger_csv(ledger)
    reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig")))
    row = next(r for r in reader if r["contract_no"] == "PO-DATE")
    assert '"contract_date": "2026-04-01"' in row["linked_sales_scopes_json"]

    xlsx_bytes = export_contract_business_ledger_xlsx(ledger)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Linked Sales Scopes"]
    headers = [c.value for c in ws[1]]
    date_col = headers.index("contract_date")
    values = [row[date_col].value for row in ws.iter_rows(min_row=2)]
    assert "2026-04-01" in values


def test_export_never_apportions_across_bridge(db_session):
    frag = _make_fragment(db_session)
    contract_a = _make_contract(db_session, frag.id, contract_no="PO-A")
    contract_b = _make_contract(db_session, frag.id, contract_no="PO-B")
    sales_contract = _make_sales_contract(db_session, frag.id, customer="Shared Customer")
    _link(db_session, contract_a, sales_contract)
    _link(db_session, contract_b, sales_contract)
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    xlsx_bytes = export_contract_business_ledger_xlsx(ledger)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    headers = [c.value for c in wb["Contract Business Ledger"][1]]
    assert "sales_invoice_amount_for_contract" not in headers
    assert "sales_receipt_amount_for_contract" not in headers
    assert "sales_total" not in headers

    scope_headers = [c.value for c in wb["Linked Sales Scopes"][1]]
    scope_rows = list(wb["Linked Sales Scopes"].iter_rows(min_row=2, values_only=True))
    sales_contract_col = scope_headers.index("sales_contract_id")
    matching = [r for r in scope_rows if r[sales_contract_col] == str(sales_contract.id)]
    assert len(matching) == 2  # once per linked procurement contract, scope facts identical


# ---------------------------------------------------------------------------
# AC — no outbound eligibility invented
# ---------------------------------------------------------------------------


def test_no_outbound_eligibility_invented(db_session):
    frag = _make_fragment(db_session)
    _make_contract(db_session, frag.id, contract_no="PO-ELIG")
    db_session.commit()

    ledger = get_contract_business_ledger(db_session)
    csv_bytes = export_contract_business_ledger_csv(ledger)
    reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig")))
    row = next(reader)
    assert row["outbound_invoice_preparation_state"] == "NOT_EVALUATED_BY_RULE"

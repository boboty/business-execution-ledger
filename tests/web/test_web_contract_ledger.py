"""Phase 2D.1-R4 — /contract-ledger web routes.

Covers GET zero-write (section 25), the CSV/XLSX download routes with
correct headers (section 42), filter parity between the page and the
exports (section 26), and the Contract 360 drill-down link (section 24).
Uses its own small synthetic DB (no SalesContract/Shipment/link data
exists in the shared Phase 2B fixture yet).
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from bel.application.procurement_sales_link import add_procurement_sales_link
from bel.application.sales_contract_facts import create_sales_contract_fact
from bel.domain.contract import Contract
from bel.domain.evidence import EvidenceDocument, EvidenceFragment, FragmentKind
from bel.domain.procurement_sales_link import ConfirmationType as LinkConfirmationType
from bel.infrastructure.persistence.database import make_engine, make_session_factory
from bel.infrastructure.persistence.models import Base
from bel.infrastructure.persistence.repositories import ContractRepository, EvidenceRepository
from bel.web.app import create_app

NOW = datetime.now(timezone.utc)


def _make_fragment(session, raw_data=None):
    doc = EvidenceDocument(
        id=uuid.uuid4(), file_name="x", sha256=uuid.uuid4().hex + uuid.uuid4().hex, source_type="t", imported_at=NOW
    )
    evidence_repo = EvidenceRepository(session)
    evidence_repo.add_document(doc)
    frag = EvidenceFragment(
        id=uuid.uuid4(),
        evidence_document_id=doc.id,
        fragment_kind=FragmentKind.MANUAL_FACT,
        sheet_name=None,
        row_number=None,
        locator_json={},
        raw_data=raw_data or {},
        created_at=NOW,
    )
    evidence_repo.add_fragment(frag)
    session.flush()
    return frag


def _make_contract(session, fragment_id, contract_no, counterparty="Supplier"):
    contract = Contract(
        id=uuid.uuid4(),
        contract_no=contract_no,
        contract_type=None,
        counterparty=counterparty,
        buyer="Our Own Entity",
        gross_amount=Decimal("1000.00"),
        currency="CNY",
        contract_date=date(2026, 1, 1),
        current_source_fragment_id=fragment_id,
        created_at=NOW,
        updated_at=NOW,
    )
    ContractRepository(session).add(contract)
    session.flush()
    return contract


def _build_ledger_db(db_path: Path) -> None:
    engine = make_engine(str(db_path))
    Base.metadata.create_all(engine)
    with make_session_factory(engine)() as session:
        frag = _make_fragment(session)
        c1 = _make_contract(session, frag.id, "PO-0001", counterparty="Alpha Supplier")
        _make_contract(session, frag.id, "PO-0002", counterparty="Beta Supplier")
        sales_contract = create_sales_contract_fact(
            session,
            our_entity="Our Own Entity",
            sales_contract_no="SC-0001",
            fields={"customer": "=cmd|'/c calc'!A0"},
            source_fragment_id=frag.id,
            created_at=NOW,
        ).sales_contract
        add_procurement_sales_link(
            session,
            procurement_contract_id=c1.id,
            sales_contract_id=sales_contract.id,
            source_fragment_id=frag.id,
            confirmation_type=LinkConfirmationType.AUTO_CONFIRMED,
            created_at=NOW,
        )
        session.commit()


@pytest.fixture
def ledger_ctx(tmp_path):
    db_path = tmp_path / "ledger-web.db"
    _build_ledger_db(db_path)
    app = create_app(str(db_path))
    client = TestClient(app)
    with app.state.session_factory() as session:
        contract_id_by_no = {c.contract_no: str(c.id) for c in ContractRepository(session).list_all()}
    return client, app, contract_id_by_no


def _db_counts(session_factory) -> dict[str, int]:
    from bel.infrastructure.persistence import models as m

    with session_factory() as session:
        counts = {}
        for name in dir(m):
            obj = getattr(m, name)
            if isinstance(obj, type) and hasattr(obj, "__tablename__"):
                counts[obj.__tablename__] = session.query(obj).count()
        return counts


def test_page_renders_and_is_zero_write(ledger_ctx):
    client, app, _ = ledger_ctx
    before = _db_counts(app.state.session_factory)
    response = client.get("/contract-ledger")
    assert response.status_code == 200
    assert "合同业务总账" in response.text
    after = _db_counts(app.state.session_factory)
    assert before == after, "GET /contract-ledger must not write a single row"


def test_page_shows_both_contracts_and_filters(ledger_ctx):
    client, app, ids = ledger_ctx
    response = client.get("/contract-ledger")
    assert "PO-0001" in response.text
    assert "PO-0002" in response.text

    filtered = client.get("/contract-ledger?supplier=Alpha")
    assert "PO-0001" in filtered.text
    assert "PO-0002" not in filtered.text


def test_contract360_drilldown_link_present(ledger_ctx):
    client, app, ids = ledger_ctx
    response = client.get("/contract-ledger")
    assert f'/contracts/{ids["PO-0001"]}' in response.text


def test_csv_export_zero_write_and_headers(ledger_ctx):
    client, app, ids = ledger_ctx
    before = _db_counts(app.state.session_factory)
    response = client.get("/contract-ledger/export.csv")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "attachment" in response.headers["content-disposition"]
    after = _db_counts(app.state.session_factory)
    assert before == after

    reader = csv.DictReader(io.StringIO(response.content.decode("utf-8-sig")))
    rows = list(reader)
    assert {r["contract_no"] for r in rows} == {"PO-0001", "PO-0002"}


def test_xlsx_export_zero_write_and_headers(ledger_ctx):
    client, app, _ = ledger_ctx
    before = _db_counts(app.state.session_factory)
    response = client.get("/contract-ledger/export.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    after = _db_counts(app.state.session_factory)
    assert before == after

    wb = load_workbook(io.BytesIO(response.content))
    assert "Contract Business Ledger" in wb.sheetnames
    assert "Contract Items" in wb.sheetnames
    assert "Shipments" in wb.sheetnames
    assert "Linked Sales Scopes" in wb.sheetnames


def test_export_filters_match_page(ledger_ctx):
    client, app, ids = ledger_ctx
    page = client.get("/contract-ledger?supplier=Alpha")
    csv_resp = client.get("/contract-ledger/export.csv?supplier=Alpha")
    reader = csv.DictReader(io.StringIO(csv_resp.content.decode("utf-8-sig")))
    csv_nos = {r["contract_no"] for r in reader}
    assert csv_nos == {"PO-0001"}
    assert "PO-0001" in page.text


def test_formula_injection_neutralized_end_to_end(ledger_ctx):
    """The seeded SalesContract customer is a formula-injection payload
    (`=cmd|'/c calc'!A0`) — the XLSX detail sheet puts it in its OWN cell
    (a genuine formula-injection surface) and must neutralize it (section
    30). Embedded inside the CSV's ``linked_sales_scopes_json`` cell it is
    NOT a formula-injection surface — that whole cell starts with ``[``,
    so a spreadsheet application never evaluates it — and is correctly
    left as literal JSON content, still round-trip parseable."""
    client, app, _ = ledger_ctx
    csv_resp = client.get("/contract-ledger/export.csv")
    csv_text = csv_resp.content.decode("utf-8-sig")
    assert "=cmd" in csv_text  # preserved verbatim inside the JSON cell
    reader = csv.DictReader(io.StringIO(csv_text))
    row = next(r for r in reader if r["contract_no"] == "PO-0001")
    assert not row["linked_sales_scopes_json"].startswith(("=", "+", "-", "@"))

    xlsx_resp = client.get("/contract-ledger/export.xlsx")
    wb = load_workbook(io.BytesIO(xlsx_resp.content))
    ws = wb["Linked Sales Scopes"]
    headers = [c.value for c in ws[1]]
    customer_col = headers.index("customer")
    values = [row[customer_col].value for row in ws.iter_rows(min_row=2)]
    assert any(v is not None and v.startswith("'=cmd") for v in values)
    for row in ws.iter_rows(min_row=2):
        assert row[customer_col].data_type != "f"


def test_html_page_escapes_injection_payload_as_text(ledger_ctx):
    """Jinja autoescapes by default — the page must render the payload as
    inert text, never break out of the DOM (defense in depth alongside
    the base app's CSP)."""
    client, app, _ = ledger_ctx
    response = client.get("/contract-ledger")
    assert "<script>" not in response.text

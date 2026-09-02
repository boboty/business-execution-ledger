"""Phase 2D.4-F2 — Exception & Task Data Product: builder + serializers.

Pure tests over constructed F1 Centers (no Session): the four exact XLSX
sheets, the unified CSV ``record_type``, source counts, canonical ids,
multi-scope completeness, unmappable retention, computed-blocker created_at
blank, period preservation, byte determinism (including across a real
wall-clock second boundary), and formula-injection neutralization on text
values (with canonical structured values untouched).
"""

from __future__ import annotations

import csv
import io
import time
import uuid
from datetime import datetime, timezone

import openpyxl
import pytest

from bel.application.exception_task_data_product import (
    CSV_HEADERS,
    XLSX_HEADERS,
    build_exception_task_data_product,
    export_exception_task_csv,
    export_exception_task_xlsx,
)
from bel.application.unresolved_work_center import (
    ComputedBlockerStatus,
    ResolutionRoute,
    ScopeType,
    SourceType,
    UnresolvedWorkCenter,
    UnresolvedWorkFilters,
    UnresolvedWorkItem,
    UnresolvedWorkScope,
)

NOW = datetime.now(timezone.utc)

XLSX_SHEET_NAMES = [
    "01_Summary",
    "02_System_Tasks",
    "03_Match_Confirmation",
    "04_Period_Close_Blockers",
]


def _item(
    source_type,
    *,
    source_id=None,
    code="CODE",
    status="OPEN",
    summary="summary text",
    created_at=None,
    scopes=(),
    invoice_id=None,
    payment_id=None,
    shipment_id=None,
    match_case_id=None,
    resolution_route=ResolutionRoute.REVIEW_ONLY,
    provenance=None,
):
    return UnresolvedWorkItem(
        source_type=source_type,
        source_id=source_id or uuid.uuid4(),
        code=code,
        status=status,
        summary=summary,
        created_at=created_at,
        scopes=scopes,
        procurement_contract_id=None,
        sales_contract_id=None,
        invoice_id=invoice_id,
        payment_id=payment_id,
        shipment_id=shipment_id,
        match_case_id=match_case_id,
        resolution_route=resolution_route,
        provenance=provenance,
    )


def _center(items, *, period=None, filters=None):
    filters = filters or UnresolvedWorkFilters(period=period)
    counts = {"total": len(items)}
    for source_type in (SourceType.TASK_EXCEPTION, SourceType.MATCH_CASE, SourceType.COMPUTED_BLOCKER):
        counts[source_type] = sum(1 for i in items if i.source_type == source_type)
    return UnresolvedWorkCenter(items=tuple(items), filters=filters, counts=counts)


def _mixed_product():
    """One of each source type — a genuine mixed Center state."""
    p1, p2 = uuid.uuid4(), uuid.uuid4()
    task = _item(
        SourceType.TASK_EXCEPTION,
        code="BusinessKeyConflict",
        summary="采购合同编号冲突",
        created_at=NOW,
        scopes=(
            UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, p1),
            UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, p2),
        ),
        provenance="bel.application.import_contract_ledger",
    )
    match = _item(
        SourceType.MATCH_CASE,
        code="M001",
        status="HUMAN_CONFIRMATION_REQUIRED",
        summary="INVOICE x 需要人工确认匹配",
        created_at=NOW,
        scopes=(UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, p1),),
        invoice_id=uuid.uuid4(),
        match_case_id=uuid.uuid4(),
        resolution_route=ResolutionRoute.CONFIRM_MATCH,
        provenance="bel.application.matching",
    )
    blocker = _item(
        SourceType.COMPUTED_BLOCKER,
        source_id="2026-03|MISSING_ACCRUAL_BASIS|aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee",
        code="MISSING_ACCRUAL_BASIS",
        status=ComputedBlockerStatus.PRESENT,
        summary="已满足成本确认条件，但缺少可确认的暂估成本依据",
        created_at=None,
        scopes=(UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, uuid.uuid4()),),
        resolution_route=ResolutionRoute.REVIEW_ONLY,
        provenance="bel.application.period_close",
    )
    return build_exception_task_data_product(_center([task, match, blocker], period="2026-03"))


# ---------------------------------------------------------------------------
# Shape — sheets, columns, counts, identity
# ---------------------------------------------------------------------------


def test_xlsx_has_exactly_four_named_sheets():
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(_mixed_product())))
    assert wb.sheetnames == XLSX_SHEET_NAMES


def test_csv_unified_record_type():
    text = export_exception_task_csv(_mixed_product()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    records = list(reader)
    assert {r["record_type"] for r in records} == {
        SourceType.TASK_EXCEPTION,
        SourceType.MATCH_CASE,
        SourceType.COMPUTED_BLOCKER,
    }
    # record_type == source_type (docs §5).
    for r in records:
        assert r["record_type"] in (SourceType.TASK_EXCEPTION, SourceType.MATCH_CASE, SourceType.COMPUTED_BLOCKER)


def test_csv_headers_are_stable_superset():
    text = export_exception_task_csv(_mixed_product()).decode("utf-8-sig")
    assert text.splitlines()[0] == ",".join(f'"{h}"' for h in CSV_HEADERS)
    assert CSV_HEADERS[0] == "record_type"
    for col in (
        "source_id",
        "code",
        "status",
        "summary",
        "created_at",
        "resolution_route",
        "provenance",
        "procurement_contract_ids",
        "sales_contract_ids",
        "invoice_id",
        "payment_id",
        "shipment_id",
        "match_case_id",
        "scopes_json",
    ):
        assert col in CSV_HEADERS


def test_xlsx_data_sheets_carry_canonical_columns():
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(_mixed_product())))
    for sheet_name in XLSX_SHEET_NAMES[1:]:
        header = [c.value for c in wb[sheet_name][1]]
        assert header == XLSX_HEADERS


def test_summary_records_counts_and_period_and_no_generated_at():
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(_mixed_product())))
    ws = wb["01_Summary"]
    field_value = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)}
    assert field_value["total"] == 3
    assert field_value["TASK_EXCEPTION"] == 1
    assert field_value["MATCH_CASE"] == 1
    assert field_value["COMPUTED_BLOCKER"] == 1
    assert field_value["period"] == "2026-03"
    for key in field_value:
        assert key not in ("generated_at", "environment", "hostname")


def test_summary_records_filter_context():
    task = _item(SourceType.TASK_EXCEPTION, code="BusinessKeyConflict", created_at=NOW)
    center = _center(
        [task],
        filters=UnresolvedWorkFilters(
            status="OPEN",
            source_type=SourceType.TASK_EXCEPTION,
            code="BusinessKeyConflict",
            procurement_contract_id=uuid.uuid4(),
            sales_contract_id=uuid.uuid4(),
            period="2026-03",
        ),
    )
    product = build_exception_task_data_product(center)
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    ws = wb["01_Summary"]
    field_value = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)}
    assert field_value["filter.status"] == "OPEN"
    assert field_value["filter.source_type"] == SourceType.TASK_EXCEPTION
    assert field_value["filter.code"] == "BusinessKeyConflict"
    assert "filter.procurement_contract_id" in field_value
    assert "filter.sales_contract_id" in field_value
    assert field_value["period"] == "2026-03"


def test_canonical_source_identity_preserved():
    product = _mixed_product()
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    task_row = next(r for r in csv.DictReader(io.StringIO(csv_text)) if r["record_type"] == SourceType.TASK_EXCEPTION)
    assert task_row["source_id"] == str(product.system_tasks[0].source_id)
    blocker_row = next(
        r for r in csv.DictReader(io.StringIO(csv_text)) if r["record_type"] == SourceType.COMPUTED_BLOCKER
    )
    # The F1 deterministic key survives byte-for-byte.
    assert blocker_row["source_id"] == "2026-03|MISSING_ACCRUAL_BASIS|aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee"


def test_multi_scope_ids_complete_not_truncated():
    p1, p2 = uuid.uuid4(), uuid.uuid4()
    task = _item(
        SourceType.TASK_EXCEPTION,
        code="BusinessKeyConflict",
        created_at=NOW,
        scopes=(
            UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, p1),
            UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, p2),
        ),
    )
    product = build_exception_task_data_product(_center([task]))
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    row = next(csv.DictReader(io.StringIO(csv_text)))
    import json

    ids = json.loads(row["procurement_contract_ids"])
    assert sorted(str(x) for x in (p1, p2)) == ids
    assert len(ids) == 2
    scopes = json.loads(row["scopes_json"])
    assert len(scopes) == 2
    assert {s["scope_id"] for s in scopes} == {str(p1), str(p2)}


def test_unmappable_task_retained_with_blank_scope_fields():
    task = _item(
        SourceType.TASK_EXCEPTION,
        code="SalesContractIdentityIncomplete",
        summary="外销合同身份信息不完整",
        created_at=NOW,
    )
    product = build_exception_task_data_product(_center([task]))
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    row = next(csv.DictReader(io.StringIO(csv_text)))
    assert row["source_id"] == str(task.source_id)
    assert row["procurement_contract_ids"] == ""
    assert row["sales_contract_ids"] == ""
    assert row["scopes_json"] == ""
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    ws = wb["02_System_Tasks"]
    header = [c.value for c in ws[1]]
    col = header.index("procurement_contract_ids") + 1
    assert ws.cell(row=2, column=col).value is None


def test_computed_blocker_created_at_blank_and_status_present():
    blocker = _item(
        SourceType.COMPUTED_BLOCKER,
        source_id="2026-03|MISSING_ACCRUAL_BASIS|abc",
        code="MISSING_ACCRUAL_BASIS",
        status=ComputedBlockerStatus.PRESENT,
        created_at=None,
    )
    product = build_exception_task_data_product(_center([blocker], period="2026-03"))
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    row = next(csv.DictReader(io.StringIO(csv_text)))
    assert row["created_at"] == ""
    assert row["status"] == ComputedBlockerStatus.PRESENT
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    ws = wb["04_Period_Close_Blockers"]
    header = [c.value for c in ws[1]]
    col = header.index("created_at") + 1
    assert ws.cell(row=2, column=col).value is None


def test_no_period_zero_blocker_rows_sheet_present():
    task = _item(SourceType.TASK_EXCEPTION, created_at=NOW)
    product = build_exception_task_data_product(_center([task], period=None))
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    assert "04_Period_Close_Blockers" in wb.sheetnames
    ws = wb["04_Period_Close_Blockers"]
    assert ws.max_row == 1  # header only
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    records = list(csv.DictReader(io.StringIO(csv_text)))
    assert all(r["record_type"] != SourceType.COMPUTED_BLOCKER for r in records)


def test_builder_accepts_only_center_and_projects_exactly():
    """The builder is a pure transformation: the exported item set is
    exactly the Center's item set, in the Center's deterministic order."""
    task = _item(SourceType.TASK_EXCEPTION, code="A", created_at=NOW)
    blocker = _item(
        SourceType.COMPUTED_BLOCKER,
        source_id="k1",
        code="MISSING_ACCRUAL_BASIS",
        status="PRESENT",
        created_at=None,
    )
    center = _center([task, blocker], period="2026-03")
    product = build_exception_task_data_product(center)
    assert [r.source_id for r in product.system_tasks] == [str(task.source_id)]
    assert [r.source_id for r in product.period_close_blockers] == ["k1"]
    assert product.match_confirmation == ()


# ---------------------------------------------------------------------------
# Determinism
# ---------------------------------------------------------------------------


def test_csv_export_byte_identical():
    product = _mixed_product()
    assert export_exception_task_csv(product) == export_exception_task_csv(product)


def test_xlsx_export_byte_identical_across_wall_clock_boundary():
    """The deterministic ZIP/core.xml normalization must keep two exports
    byte-identical across a REAL second boundary, not just in-process."""
    product = _mixed_product()
    first = export_exception_task_xlsx(product)
    time.sleep(1.5)
    second = export_exception_task_xlsx(product)
    assert first == second


# ---------------------------------------------------------------------------
# Formula injection safety
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("dangerous", ["=CMD()", "+ABC", "-ABC", "@ABC"])
def test_formula_injection_neutralized_in_csv_and_xlsx(dangerous):
    task = _item(SourceType.TASK_EXCEPTION, code="BusinessKeyConflict", summary=dangerous, created_at=NOW)
    product = build_exception_task_data_product(_center([task]))

    csv_bytes = export_exception_task_csv(product)
    expected = dangerous.encode("utf-8")
    assert ("'" + dangerous).encode("utf-8") in csv_bytes
    assert expected in csv_bytes  # the literal text survives, prefixed

    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    ws = wb["02_System_Tasks"]
    header = [c.value for c in ws[1]]
    col = header.index("summary") + 1
    assert str(ws.cell(row=2, column=col).value).startswith("'")


def test_canonical_structured_values_not_corrupted_by_formula_guard():
    p = uuid.uuid4()
    task = _item(
        SourceType.TASK_EXCEPTION,
        source_id=p,
        code="BusinessKeyConflict",
        status="OPEN",
        created_at=NOW,
        scopes=(UnresolvedWorkScope(ScopeType.PROCUREMENT_CONTRACT, p),),
    )
    product = build_exception_task_data_product(_center([task]))
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    row = next(csv.DictReader(io.StringIO(csv_text)))
    assert row["source_id"] == str(p)  # exact, no prefix
    assert row["code"] == "BusinessKeyConflict"
    assert row["status"] == "OPEN"
    wb = openpyxl.load_workbook(io.BytesIO(export_exception_task_xlsx(product)))
    ws = wb["02_System_Tasks"]
    header = [c.value for c in ws[1]]
    assert ws.cell(row=2, column=header.index("source_id") + 1).value == str(p)
    assert ws.cell(row=2, column=header.index("code") + 1).value == "BusinessKeyConflict"


def test_no_advisory_or_missing_contract_gross_amount_codes():
    """The Data Product exports exactly what the F1 Center contains — a
    constructed Center cannot carry an advisory/MISSING_CONTRACT_GROSS_AMOUNT
    item, and no such code is ever synthesized."""
    product = _mixed_product()
    csv_text = export_exception_task_csv(product).decode("utf-8-sig")
    records = list(csv.DictReader(io.StringIO(csv_text)))
    codes = {r["code"] for r in records}
    assert "MISSING_CONTRACT_GROSS_AMOUNT" not in codes
    assert not (codes & {"SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED", "SALES_INVOICE_AMOUNT_DEVIATION"})


def test_xlsx_package_metadata_fully_pinned():
    """G0 repair #2 (Blocker B): the produced XLSX ZIP must carry NO
    wall-clock metadata — every entry date_time and both core.xml
    timestamps are fixed (structural guarantee behind the wall-clock byte
    test above)."""
    from tests.xlsx_assertions import assert_xlsx_package_metadata_fixed

    product = _mixed_product()
    assert_xlsx_package_metadata_fixed(export_exception_task_xlsx(product))

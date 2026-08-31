"""Period Close Data Product tests (Phase 2D.2). Uses the same synthetic
Phase 2B close fixture as the Workbench tests — the Data Product is a
projection of the SAME Workbench, so it must show the SAME business
outcomes (parity requirement, docs/PHASE2D2-DECISIONS.md section 10).
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone

import openpyxl

from bel.application.period_close import build_period_close_preview
from bel.application.period_close_export import (
    RECORD_TYPE_ACCRUAL_REQUIRED,
    RECORD_TYPE_ACTUAL_DIFFERENCE,
    RECORD_TYPE_BLOCKER,
    RECORD_TYPE_CONTRACT_LEVEL_CANDIDATE,
    RECORD_TYPE_PRIOR_ACCRUAL_REVERSAL,
    build_period_close_data_product,
    export_period_close_csv,
    export_period_close_xlsx,
)
from bel.application.period_close_workbench import get_period_close_workbench
from bel.domain.matching import (
    AllocationMatchMethod,
    ConfirmationType,
    InvoiceAllocation,
    MatchCase,
    MatchCaseStatus,
    MatchMethod,
)
from bel.infrastructure.persistence.models import (
    AccrualModel,
    AccrualReversalModel,
    ContractItemModel,
    ContractModel,
    InvoiceItemAllocationModel,
)
from bel.infrastructure.persistence.repositories import (
    ContractRepository,
    InvoiceAllocationRepository,
    InvoiceRepository,
    MatchCaseRepository,
)
from fixtures.synthetic.phase2b_close import CLOSE_PERIOD

from bel.application.import_close_facts import import_close_facts
from bel.application.import_contract_ledger import import_contract_ledger
from bel.application.import_invoices import import_invoices
from bel.domain.invoice import InvoiceDirection

EXPECTED_SHEETS = [
    "01_Summary",
    "02_Accrual_Required",
    "03_Prior_Accrual_Reversal",
    "04_Actual_Difference",
    "05_Contract_Level_Candidate",
    "06_Blockers",
]


def _confirm_invoice_contract(session, invoice, contract) -> None:
    now = datetime.now(timezone.utc)
    match_case = MatchCase(
        id=uuid.uuid4(),
        subject_type="INVOICE",
        subject_id=invoice.id,
        status=MatchCaseStatus.AUTO_CONFIRMED,
        match_method=MatchMethod.M001,
        created_at=now,
        resolved_at=now,
    )
    MatchCaseRepository(session).add(match_case)
    session.flush()
    InvoiceAllocationRepository(session).add(
        InvoiceAllocation(
            id=uuid.uuid4(),
            invoice_id=invoice.id,
            contract_id=contract.id,
            match_case_id=match_case.id,
            allocated_gross_amount=invoice.gross_amount,
            match_method=AllocationMatchMethod.EXACT_COUNTERPARTY_AMOUNT_UNIQUE,
            confirmation_type=ConfirmationType.AUTO_CONFIRMED,
            created_at=now,
        )
    )


def _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path) -> None:
    import_contract_ledger(db_session, phase2b_ledger_path)
    import_invoices(db_session, phase2b_invoices_path, InvoiceDirection.PURCHASE)
    for external_key, contract_no in [
        ("DIGITAL-CLOSE-001", "PO-CLOSE-001"),
        ("DIGITAL-CLOSE-002", "PO-CLOSE-002"),
        ("DIGITAL-CLOSE-005", "PO-CLOSE-005"),
        ("DIGITAL-CLOSE-006", "PO-CLOSE-006"),
    ]:
        invoice = InvoiceRepository(db_session).find_by_external_key(external_key)
        contract = next(c for c in ContractRepository(db_session).list_all() if c.contract_no == contract_no)
        _confirm_invoice_contract(db_session, invoice, contract)
    db_session.commit()
    import_close_facts(db_session, phase2b_close_facts_path)


def test_data_product_parity_with_preview(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)

    preview = build_period_close_preview(db_session, CLOSE_PERIOD)
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)

    assert len(product.accrual_required) == len(preview.new_accrual_requirements) > 0
    assert len(product.prior_accrual_reversal) == len(preview.prior_accrual_reversals) > 0
    assert len(product.actual_difference) == len(preview.accrual_actual_differences) > 0
    assert len(product.contract_level_candidate) == len(preview.contract_level_candidates) > 0
    assert len(product.blocker) == len(preview.blockers) > 0

    # every logical record type appears in the unified row set exactly
    # once per underlying decision — no export-only decision, none lost.
    record_types = {row.record_type for row in product.all_rows}
    assert record_types == {
        RECORD_TYPE_ACCRUAL_REQUIRED,
        RECORD_TYPE_PRIOR_ACCRUAL_REVERSAL,
        RECORD_TYPE_ACTUAL_DIFFERENCE,
        RECORD_TYPE_CONTRACT_LEVEL_CANDIDATE,
        RECORD_TYPE_BLOCKER,
    }
    assert len(product.all_rows) == (
        len(preview.new_accrual_requirements)
        + len(preview.prior_accrual_reversals)
        + len(preview.accrual_actual_differences)
        + len(preview.contract_level_candidates)
        + len(preview.blockers)
    )


def test_xlsx_has_exact_six_sheets_in_order(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)

    xlsx_bytes = export_period_close_xlsx(product)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_xlsx_summary_sheet_identifies_period(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)

    xlsx_bytes = export_period_close_xlsx(product)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    summary_rows = {row[0]: row[1] for row in wb["01_Summary"].iter_rows(min_row=2, values_only=True)}
    assert summary_rows["period"] == CLOSE_PERIOD
    assert summary_rows["period_end"] == "2031-03-31"
    assert summary_rows["blockers"] == 2


def test_xlsx_decision_sheets_no_business_status_invented(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    """Never invent CLOSED/POSTED/APPROVED anywhere in the export
    (docs/PHASE2D2-DECISIONS.md section 4.1's HARD rule)."""
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)

    xlsx_bytes = export_period_close_xlsx(product)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    for sheet_name in EXPECTED_SHEETS:
        for row in wb[sheet_name].iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    assert cell not in {"CLOSED", "POSTED", "APPROVED"}


def test_csv_unified_long_table_all_record_types(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)

    csv_bytes = export_period_close_csv(product)
    text = csv_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    assert reader.fieldnames[:3] == ["period", "period_end", "record_type"]

    record_types = {row["record_type"] for row in rows}
    assert record_types == {
        RECORD_TYPE_ACCRUAL_REQUIRED,
        RECORD_TYPE_PRIOR_ACCRUAL_REVERSAL,
        RECORD_TYPE_ACTUAL_DIFFERENCE,
        RECORD_TYPE_CONTRACT_LEVEL_CANDIDATE,
        RECORD_TYPE_BLOCKER,
    }
    assert len(rows) == len(product.all_rows)

    # a candidate row has no source_item_key (contract-level, not item-level) —
    # blank, never fabricated.
    candidate_rows = [r for r in rows if r["record_type"] == RECORD_TYPE_CONTRACT_LEVEL_CANDIDATE]
    assert candidate_rows
    for row in candidate_rows:
        assert row["source_item_key"] == ""
        assert row["quantity"] == ""


def test_evidence_trace_is_decision_provenance_and_blocker_context_is_separate(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)

    assert product.prior_accrual_reversal
    for row in product.prior_accrual_reversal:
        # genuine Decision -> Fact -> Evidence provenance stays in
        # evidence_trace, unchanged
        assert row.evidence_trace
        assert "HISTORICAL_ACCRUAL" in row.evidence_trace
        assert row.blocker_context is None

    assert product.blocker
    for row in product.blocker:
        # BlockerContext is current business context, NOT Evidence
        # provenance: it renders under blocker_context, never under
        # evidence_trace — and no Fact -> Evidence chain is invented for
        # blockers (none exists).
        assert isinstance(row.blocker_context, str)
        assert row.blocker_context
        assert row.evidence_trace is None

    # the Blocker sheet and the unified CSV schema expose the context
    # under its own column, not under evidence_trace
    wb = openpyxl.load_workbook(io.BytesIO(export_period_close_xlsx(product)))
    blocker_headers = [c.value for c in wb["06_Blockers"][1]]
    assert "blocker_context" in blocker_headers
    assert "evidence_trace" not in blocker_headers

    reader = csv.DictReader(io.StringIO(export_period_close_csv(product).decode("utf-8-sig")))
    assert "blocker_context" in reader.fieldnames
    assert "evidence_trace" in reader.fieldnames
    for row in reader:
        if row["record_type"] == RECORD_TYPE_BLOCKER:
            assert row["evidence_trace"] == ""
            assert row["blocker_context"] != ""


def test_export_is_read_only(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)

    def _counts():
        return {
            "contracts": db_session.query(ContractModel).count(),
            "contract_items": db_session.query(ContractItemModel).count(),
            "accruals": db_session.query(AccrualModel).count(),
            "accrual_reversals": db_session.query(AccrualReversalModel).count(),
            "invoice_item_allocations": db_session.query(InvoiceItemAllocationModel).count(),
        }

    before = _counts()
    workbench = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product = build_period_close_data_product(workbench)
    export_period_close_xlsx(product)
    export_period_close_csv(product)
    after = _counts()
    assert before == after


def test_semantic_rerun_is_deterministic(
    db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path
):
    _seed_full_fixture(db_session, phase2b_ledger_path, phase2b_invoices_path, phase2b_close_facts_path)

    workbench1 = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product1 = build_period_close_data_product(workbench1)
    workbench2 = get_period_close_workbench(db_session, CLOSE_PERIOD)
    product2 = build_period_close_data_product(workbench2)

    assert product1 == product2
    assert export_period_close_csv(product1) == export_period_close_csv(product2)

    wb1 = openpyxl.load_workbook(io.BytesIO(export_period_close_xlsx(product1)))
    wb2 = openpyxl.load_workbook(io.BytesIO(export_period_close_xlsx(product2)))
    for sheet_name in EXPECTED_SHEETS:
        rows1 = list(wb1[sheet_name].iter_rows(values_only=True))
        rows2 = list(wb2[sheet_name].iter_rows(values_only=True))
        assert rows1 == rows2


def test_empty_database_produces_empty_but_valid_product(db_session):
    workbench = get_period_close_workbench(db_session, "2031-01")
    product = build_period_close_data_product(workbench)

    assert product.all_rows == ()
    assert product.summary == {
        "period": "2031-01",
        "prior_accrual_reversals": 0,
        "new_accrual_requirements": 0,
        "contract_level_candidates": 0,
        "accrual_actual_differences": 0,
        "blockers": 0,
    }

    csv_bytes = export_period_close_csv(product)
    text = csv_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    assert list(reader) == []

    xlsx_bytes = export_period_close_xlsx(product)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == EXPECTED_SHEETS
    for sheet_name in EXPECTED_SHEETS[1:]:
        assert wb[sheet_name].max_row == 1  # header only

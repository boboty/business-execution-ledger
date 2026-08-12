"""Excel adapter for the 进项发票 (purchase invoice) ledger export.

Structural quirk this adapter exists to handle (spec section 7): the
first row of a multi-item invoice carries BOTH the invoice header AND
item 1; subsequent items are continuation rows carrying only item
fields (invoice_no/seller/date are blank on those rows — never
forward-filled). A header row is any row with digital_invoice_no or
invoice_no present; everything until the next header row belongs to
that invoice.

Parsing policy mirrors adapters/excel/contract_ledger.py: every cell is
preserved into raw_data verbatim (datetime -> ISO string is the only
transformation, for JSON-storability); direction is never guessed here
— it comes from the caller (CLI --direction flag).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

SHEET_NAME = "sheet1"
HEADER_ROW = 3
DATA_START_ROW = 4
BUYER_CELL = "A2"

INVOICE_TYPE_HEADER = "发票票种"
ISSUE_DATE_HEADER = "开票日期"
INVOICE_NO_HEADER = "发票号码"
DIGITAL_INVOICE_NO_HEADER = "数电发票号码"
SELLER_HEADER = "销方名称"
INVOICE_STATUS_HEADER = "发票状态"
INVOICE_NET_HEADER = "发票金额"
INVOICE_TAX_HEADER = "发票税额"
INVOICE_GROSS_HEADER = "发票价税合计"

ITEM_PRODUCT_HEADER = "商品名称（明细）"
ITEM_SPEC_HEADER = "规格型号（明细）"
ITEM_UNIT_HEADER = "单位（明细）"
ITEM_QTY_HEADER = "数量（明细）"
ITEM_UNIT_PRICE_HEADER = "单价（明细）"
ITEM_NET_HEADER = "金额（明细）"
ITEM_TAX_RATE_HEADER = "税率（%）（明细）"
ITEM_TAX_HEADER = "税额（明细）"
ITEM_GROSS_HEADER = "价税合计（明细）"


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _none_if_blank(value: Any) -> Any:
    return None if _is_blank(value) else value


def _serialize_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _decimal_required(value: Any, context: str) -> Decimal:
    if _is_blank(value):
        raise ValueError(f"{context}: expected a numeric value, got blank")
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"{context}: cannot parse {value!r} as a decimal amount") from exc


def _decimal_or_zero(value: Any) -> Decimal:
    """Blank tax fields mean tax-exempt (0%), not missing data — see
    docs/PHASE2A-DECISIONS.md. Acceptance checks require blank 税额/税率
    to pair with net_amount == gross_amount."""
    if _is_blank(value):
        return Decimal("0")
    return Decimal(str(value))


def _decimal_or_none(value: Any) -> Decimal | None:
    if _is_blank(value):
        return None
    return Decimal(str(value))


def _parse_date_or_none(value: Any) -> date | None:
    if _is_blank(value):
        return None
    return date.fromisoformat(str(value).strip())


@dataclass
class ParsedInvoiceRow:
    row_number: int
    raw_data: dict[str, Any]
    is_invoice_header: bool

    product_name: str | None
    specification: str | None
    unit: str | None
    quantity: Decimal | None
    unit_price: Decimal | None
    item_net_amount: Decimal
    tax_rate: Decimal | None
    item_tax_amount: Decimal
    item_gross_amount: Decimal

    # Populated only when is_invoice_header is True.
    invoice_type: str | None = None
    invoice_no: str | None = None
    digital_invoice_no: str | None = None
    issue_date: date | None = None
    seller: str | None = None
    invoice_status: str | None = None
    invoice_net_amount: Decimal | None = None
    invoice_tax_amount: Decimal | None = None
    invoice_gross_amount: Decimal | None = None


@dataclass
class InvoiceGroup:
    header: ParsedInvoiceRow
    item_rows: list[ParsedInvoiceRow]  # includes the header row itself as item 1


@dataclass
class ParsedInvoiceWorkbook:
    sheet_name: str
    buyer: str | None
    rows: list[ParsedInvoiceRow]
    groups: list[InvoiceGroup]


def parse_invoice_ledger(path: Path) -> ParsedInvoiceWorkbook:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"expected sheet {SHEET_NAME!r} not found in workbook; sheets present: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[HEADER_ROW]]
    if any(h is None for h in headers):
        raise ValueError(f"sheet {SHEET_NAME!r} header row {HEADER_ROW} has one or more empty column headers: {headers}")

    buyer = _none_if_blank(ws[BUYER_CELL].value)
    if isinstance(buyer, str):
        buyer = buyer.strip()

    rows: list[ParsedInvoiceRow] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        raw_data: dict[str, Any] = {}
        for ci, header in enumerate(headers, start=1):
            raw_data[header] = _serialize_cell(ws.cell(row=r, column=ci).value)

        digital_invoice_no = raw_data.get(DIGITAL_INVOICE_NO_HEADER)
        invoice_no = raw_data.get(INVOICE_NO_HEADER)
        is_header = not _is_blank(digital_invoice_no) or not _is_blank(invoice_no)

        row = ParsedInvoiceRow(
            row_number=r,
            raw_data=raw_data,
            is_invoice_header=is_header,
            product_name=_none_if_blank(raw_data.get(ITEM_PRODUCT_HEADER)),
            specification=_none_if_blank(raw_data.get(ITEM_SPEC_HEADER)),
            unit=_none_if_blank(raw_data.get(ITEM_UNIT_HEADER)),
            quantity=_decimal_or_none(raw_data.get(ITEM_QTY_HEADER)),
            unit_price=_decimal_or_none(raw_data.get(ITEM_UNIT_PRICE_HEADER)),
            item_net_amount=_decimal_required(raw_data.get(ITEM_NET_HEADER), f"row {r} {ITEM_NET_HEADER}"),
            tax_rate=_decimal_or_none(raw_data.get(ITEM_TAX_RATE_HEADER)),
            item_tax_amount=_decimal_or_zero(raw_data.get(ITEM_TAX_HEADER)),
            item_gross_amount=_decimal_required(raw_data.get(ITEM_GROSS_HEADER), f"row {r} {ITEM_GROSS_HEADER}"),
        )

        if is_header:
            row.invoice_type = _none_if_blank(raw_data.get(INVOICE_TYPE_HEADER))
            row.invoice_no = _none_if_blank(invoice_no)
            row.digital_invoice_no = _none_if_blank(digital_invoice_no)
            row.issue_date = _parse_date_or_none(raw_data.get(ISSUE_DATE_HEADER))
            row.seller = _none_if_blank(raw_data.get(SELLER_HEADER))
            row.invoice_status = _none_if_blank(raw_data.get(INVOICE_STATUS_HEADER))
            row.invoice_net_amount = _decimal_or_zero(raw_data.get(INVOICE_NET_HEADER))
            row.invoice_tax_amount = _decimal_or_zero(raw_data.get(INVOICE_TAX_HEADER))
            row.invoice_gross_amount = _decimal_or_zero(raw_data.get(INVOICE_GROSS_HEADER))

        rows.append(row)

    groups: list[InvoiceGroup] = []
    current: list[ParsedInvoiceRow] | None = None
    for row in rows:
        if row.is_invoice_header:
            if current is not None:
                groups.append(InvoiceGroup(header=current[0], item_rows=current))
            current = [row]
        else:
            if current is None:
                raise ValueError(
                    f"row {row.row_number}: continuation row (no invoice_no/digital_invoice_no) with no "
                    "preceding invoice header row"
                )
            current.append(row)
    if current is not None:
        groups.append(InvoiceGroup(header=current[0], item_rows=current))

    return ParsedInvoiceWorkbook(sheet_name=SHEET_NAME, buyer=buyer, rows=rows, groups=groups)

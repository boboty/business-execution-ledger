"""Excel adapter for the 合同台账 (contract ledger) workbook.

Parsing policy (see docs/PHASE1-DECISIONS.md for the full rationale):
  - Every cell in the primary sheet is preserved into raw_data verbatim.
    The only transformation applied is making non-JSON-serializable
    Python values (datetime/date) into ISO strings — that is lossless
    serialization, not interpretation.
  - Only four columns are ever promoted into canonical Contract fields:
    合同编码 / 卖方 / 买方 / 金额. Every other column (是否付款, 进项发票
    入账月, 退税到账日期, 备注, ...) is Evidence-only in Phase 1 — see
    docs/RULES.md and section 6/8 of the Phase 1 task spec.
  - No inference: contract_date is never derived from contract_no or any
    other column. Amounts are parsed as Decimal, never float.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from bel.adapters.common import compute_sha256

PRIMARY_SHEET_NAME = "报关出口购销合同"
HEADER_ROW = 2
DATA_START_ROW = 3

CONTRACT_NO_HEADER = "合同编码"
COUNTERPARTY_HEADER = "卖方"
BUYER_HEADER = "买方"
GROSS_AMOUNT_HEADER = "金额"


def _serialize_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


@dataclass
class ParsedRow:
    row_number: int
    sheet_name: str
    raw_data: dict[str, Any]
    is_business_row: bool
    contract_no: str | None
    counterparty: str | None
    buyer: str | None
    gross_amount: Decimal | None


@dataclass
class ParsedWorkbook:
    sheet_names: list[str]
    primary_sheet: str
    primary_sheet_columns: int
    rows: list[ParsedRow]

    @property
    def business_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.is_business_row]

    @property
    def blank_trailing_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if not r.is_business_row]


def parse_contract_ledger(path: Path) -> ParsedWorkbook:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = list(wb.sheetnames)
    if PRIMARY_SHEET_NAME not in sheet_names:
        raise ValueError(
            f"expected primary sheet {PRIMARY_SHEET_NAME!r} not found in workbook; sheets present: {sheet_names}"
        )

    ws = wb[PRIMARY_SHEET_NAME]
    headers = [c.value for c in ws[HEADER_ROW]]
    if any(h is None for h in headers):
        raise ValueError(
            f"primary sheet {PRIMARY_SHEET_NAME!r} header row {HEADER_ROW} has one or more empty column headers: {headers}"
        )

    rows: list[ParsedRow] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        raw_data: dict[str, Any] = {}
        for ci, header in enumerate(headers, start=1):
            raw_data[header] = _serialize_cell(ws.cell(row=r, column=ci).value)

        contract_no_raw = raw_data.get(CONTRACT_NO_HEADER)
        is_business_row = contract_no_raw is not None and str(contract_no_raw).strip() != ""

        contract_no = None
        counterparty = None
        buyer = None
        gross_amount = None

        if is_business_row:
            contract_no = str(contract_no_raw).strip()
            counterparty = raw_data.get(COUNTERPARTY_HEADER)
            buyer = raw_data.get(BUYER_HEADER)

            amount_value = raw_data.get(GROSS_AMOUNT_HEADER)
            if amount_value is None:
                raise ValueError(f"row {r}: business row (contract_no={contract_no!r}) has no {GROSS_AMOUNT_HEADER!r}")
            try:
                gross_amount = Decimal(str(amount_value))
            except InvalidOperation as exc:
                raise ValueError(
                    f"row {r}: cannot parse {GROSS_AMOUNT_HEADER!r}={amount_value!r} as a decimal amount"
                ) from exc

        rows.append(
            ParsedRow(
                row_number=r,
                sheet_name=PRIMARY_SHEET_NAME,
                raw_data=raw_data,
                is_business_row=is_business_row,
                contract_no=contract_no,
                counterparty=counterparty,
                buyer=buyer,
                gross_amount=gross_amount,
            )
        )

    return ParsedWorkbook(
        sheet_names=sheet_names,
        primary_sheet=PRIMARY_SHEET_NAME,
        primary_sheet_columns=len(headers),
        rows=rows,
    )

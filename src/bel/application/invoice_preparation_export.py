"""Invoice Preparation Data Product (Phase 2D.3-F2b).

Turns the read-only ``InvoicePreparationWorkbench`` into a deliverable,
traceable, reproducible Data Product. The ONE Application-layer path both
Web and CLI call:

    get_invoice_preparation_workbench(session)
        -> build_invoice_preparation_data_product(workbench)   (this module)
        -> export_invoice_preparation_xlsx() / export_invoice_preparation_csv()

The neutral Data Product DTO is built ONLY from the Workbench — never
from a Session/repositories and never by re-running a business rule. The
F1 comparison / cardinality / currency-safety / follow-up outcomes are
flattened exactly as the F1 reports produced them: no sum, no
apportionment, no currency inference, no new advisory.

The three semantic layers stay explicit (FACT / COMPARISON / ATTENTION);
they are never flattened into READY / BLOCKED / ELIGIBLE. Four record
types: SALES_PREPARATION, SALES_ATTENTION, SUPPLIER_REQUEST,
SUPPLIER_ATTENTION. Attention keeps three distinguishable categories:
UNRESOLVED_WORK (existing Task/Exception context), INCOMPLETE_ASSOCIATION
(an allocation whose referenced base Fact is missing or mismatched —
NEVER counted as a confirmed Fact, F2a repair), and MANAGEMENT_ADVISORY
(recomputed F1 management signals such as the P09 follow-up).

Canonical machine-readable codes (MATCH / DEVIATION / NOT_COMPARABLE_* /
SALES_INVOICE_AMOUNT_DEVIATION / SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED
...) are preserved; a concise business-facing message is carried alongside
them, but Chinese display text is never the only machine contract.

Strictly read-only and reproducible: repeated export over unchanged
canonical state is byte-stable (deterministic row ordering and columns,
no generated_at, no random ids, no environment-specific values), and
export performs zero database writes. Decimal amounts are written as
numeric XLSX cells; a missing Fact is a blank cell, never 0 / a guessed
currency / an invented "N/A".
"""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from bel.application.invoice_preparation_workbench import InvoicePreparationWorkbench
from bel.domain.invoice import InvoiceDirection
from bel.domain.payment import PaymentDirection

# ---------------------------------------------------------------------------
# Vocabulary — record types and attention categories (the neutral contract)
# ---------------------------------------------------------------------------

RECORD_TYPE_SALES_PREPARATION = "SALES_PREPARATION"
RECORD_TYPE_SALES_ATTENTION = "SALES_ATTENTION"
RECORD_TYPE_SUPPLIER_REQUEST = "SUPPLIER_REQUEST"
RECORD_TYPE_SUPPLIER_ATTENTION = "SUPPLIER_ATTENTION"

ATTENTION_CATEGORY_UNRESOLVED_WORK = "UNRESOLVED_WORK"
ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION = "INCOMPLETE_ASSOCIATION"
ATTENTION_CATEGORY_MANAGEMENT_ADVISORY = "MANAGEMENT_ADVISORY"

_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# Fixed XLSX created/modified timestamp so repeated exports of identical
# state are byte-identical regardless of wall-clock time (openpyxl would
# otherwise stamp the current time into docProps/core.xml).
_FIXED_XLSX_DATETIME = datetime(1980, 1, 1, 0, 0, 0)


# ---------------------------------------------------------------------------
# Business-facing messages — carried ALONGSIDE the canonical codes, never
# replacing them as the machine contract.
# ---------------------------------------------------------------------------

SALES_COMPARISON_MESSAGES = {
    "MATCH": "金额核对一致",
    "DEVIATION": "金额存在偏差，建议复核",
    "NOT_COMPARABLE_MISSING_FACT": "当前信息不足，暂无法核对",
    "NOT_COMPARABLE_CURRENCY_MISMATCH": "币种不同，暂不直接比较金额",
    "NOT_COMPARABLE_AMBIGUOUS_SCOPE": "对应范围不唯一，暂无法自动核对",
}

SUPPLIER_AMOUNT_CHECK_MESSAGES = {
    "MATCH": "已有发票与参考信息一致",
    "DEVIATION": "存在金额偏差，建议复核",
    "NOT_COMPARABLE_MISSING_FACT": "当前信息不足或范围无法直接比较",
    "NOT_COMPARABLE_CURRENCY_MISMATCH": "币种不同，暂不直接比较金额",
}

SUPPLIER_ITEM_NAME_CHECK_MESSAGES = {
    "MATCH": "商品名称与合同确认名称一致",
    "DEVIATION": "商品名称与合同确认名称不一致，建议复核",
    "NOT_COMPARABLE_MISSING_FACT": "当前信息不足，暂无法核对",
}

ADVISORY_MESSAGES = {
    "SALES_INVOICE_AMOUNT_DEVIATION": "销项发票金额与合同/报关金额存在偏差，建议复核",
    "SALES_INVOICE_CURRENCY_DEVIATION": "销项发票币种与合同/报关币种不一致，暂不直接比较金额，建议复核",
    "SUPPLIER_INVOICE_FOLLOW_UP_RECOMMENDED": "已付款，尚未收到对应进项发票，建议催供应商开票",
    "PURCHASE_INVOICE_AMOUNT_DEVIATION": "采购发票金额与合同参考金额存在偏差，建议复核",
    "PURCHASE_INVOICE_CURRENCY_DEVIATION": "采购发票币种与合同参考币种不一致，暂不直接比较金额，建议复核",
    "PURCHASE_INVOICE_PRODUCT_NAME_DEVIATION": "商品名称与合同确认名称不一致，建议复核",
    "MULTIPLE_PURCHASE_INVOICES_ON_CONTRACT": "一个采购合同关联多张已确认采购发票，建议复核",
    "PURCHASE_INVOICE_SPANS_MULTIPLE_CONTRACTS": "一张采购发票关联多个采购合同，建议复核",
}

INCOMPLETE_ASSOCIATION_KIND_LABELS = {
    "SALES_INVOICE": "销项发票关联",
    "SALES_RECEIPT": "收款关联",
    "PURCHASE_INVOICE": "采购发票关联",
    "OUT_PAYMENT": "付款关联",
}


# ---------------------------------------------------------------------------
# Neutral Data Product DTO — one superset row schema, all four record types.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class InvoicePreparationExportRow:
    """One presentation-neutral record. Every Data Product row — CSV or
    any XLSX sheet — is a projection of this same shape; a field that does
    not apply to this ``record_type`` stays ``None``. Amounts are the
    Facts/values the Workbench already carried; ``None`` means the Fact is
    absent (blank cell), never a manufactured zero or default."""

    record_type: str
    # Scope identity — canonical business identifiers for navigation /
    # reconciliation (no Evidence trace is fabricated; the Workbench DTO
    # does not carry it).
    sales_contract_id: str | None = None
    sales_contract_no: str | None = None
    procurement_contract_id: str | None = None
    contract_no: str | None = None
    # Facts
    our_entity: str | None = None
    customer: str | None = None
    supplier: str | None = None
    contract_gross_amount: Decimal | None = None
    contract_currency: str | None = None
    expected_invoice_amount: Decimal | None = None
    expected_invoice_currency: str | None = None
    linked_procurement_contract_count: int | None = None
    confirmed_sales_invoice_count: int | None = None
    confirmed_receipt_count: int | None = None
    confirmed_purchase_invoice_count: int | None = None
    confirmed_out_payment_count: int | None = None
    # Comparison (sales IP-S02 three-way; the singular supplier amount check
    # is exposed separately below)
    comparison_outcome: str | None = None
    comparison_message: str | None = None
    sales_contract_amount: Decimal | None = None
    sales_contract_currency: str | None = None
    declared_amount: Decimal | None = None
    declared_currency: str | None = None
    sales_invoice_amount: Decimal | None = None
    sales_invoice_currency: str | None = None
    # The F1f comparison's canonical trace identifiers — the resolved
    # Shipment/Export Fact and confirmed SALES Invoice Fact the comparison
    # was scoped to (None when missing or ambiguous). Never reconstructed.
    comparison_shipment_id: str | None = None
    comparison_sales_invoice_id: str | None = None
    # Supplier P02 amount-check outcome (naturally 0 or 1 per contract),
    # explicit check-count fields, and the full deterministic JSON
    # serialization of ALL supplier checks.
    supplier_amount_check_outcome: str | None = None
    supplier_amount_check_count: int | None = None
    supplier_item_name_check_count: int | None = None
    supplier_item_name_deviation_count: int | None = None
    supplier_checks_json: str | None = None
    # Attention — the canonical scope identity is carried on every
    # attention row (contract_no / sales_contract_no alone are NOT unique),
    # and management-advisory rows preserve their full canonical related-id
    # sets as deterministic JSON arrays (never a truncated first id).
    attention_category: str | None = None
    attention_code: str | None = None
    attention_message: str | None = None
    related_invoice_ids: str | None = None
    related_contract_ids: str | None = None
    related_invoice_item_ids: str | None = None
    related_shipment_ids: str | None = None
    source_id: str | None = None
    allocated_amount: Decimal | None = None


@dataclass(frozen=True)
class InvoicePreparationDataProduct:
    """The neutral Data Product DTO. XLSX/CSV serialization reads only
    this — never the Workbench or raw repositories directly."""

    summary: dict[str, int]
    sales_preparation: tuple[InvoicePreparationExportRow, ...]
    sales_attention: tuple[InvoicePreparationExportRow, ...]
    supplier_request: tuple[InvoicePreparationExportRow, ...]
    supplier_attention: tuple[InvoicePreparationExportRow, ...]

    @property
    def all_rows(self) -> tuple[InvoicePreparationExportRow, ...]:
        return (
            self.sales_preparation
            + self.sales_attention
            + self.supplier_request
            + self.supplier_attention
        )


# ---------------------------------------------------------------------------
# Confirmed-Fact boundary (F2a repair carried into the Data Product): an
# allocation/association is NEVER proof that the referenced Fact exists.
# ---------------------------------------------------------------------------


def _confirmed_sales_invoices(scope) -> tuple:
    return tuple(
        e
        for e in scope.invoice_allocations
        if e.invoice is not None and e.invoice.direction == InvoiceDirection.SALES
    )


def _incomplete_sales_invoices(scope) -> tuple:
    return tuple(
        e
        for e in scope.invoice_allocations
        if not (e.invoice is not None and e.invoice.direction == InvoiceDirection.SALES)
    )


def _confirmed_in_receipts(scope) -> tuple:
    return tuple(
        e
        for e in scope.payment_allocations
        if e.payment is not None and e.payment.direction == PaymentDirection.IN
    )


def _incomplete_sales_receipts(scope) -> tuple:
    return tuple(
        e
        for e in scope.payment_allocations
        if not (e.payment is not None and e.payment.direction == PaymentDirection.IN)
    )


def _confirmed_purchase_invoices(scope) -> tuple:
    return tuple(
        e
        for e in scope.invoice_allocations
        if e.invoice is not None and e.invoice.direction == InvoiceDirection.PURCHASE
    )


def _incomplete_purchase_invoices(scope) -> tuple:
    return tuple(
        e
        for e in scope.invoice_allocations
        if not (e.invoice is not None and e.invoice.direction == InvoiceDirection.PURCHASE)
    )


def _confirmed_out_payments(scope) -> tuple:
    return tuple(
        e
        for e in scope.payment_allocations
        if e.payment is not None and e.payment.direction == PaymentDirection.OUT
    )


def _incomplete_out_payments(scope) -> tuple:
    return tuple(
        e
        for e in scope.payment_allocations
        if not (e.payment is not None and e.payment.direction == PaymentDirection.OUT)
    )


# ---------------------------------------------------------------------------
# Row builders — pure projection of the Workbench's context + F1 decisions.
# ---------------------------------------------------------------------------


def _sales_preparation_row(scope, decision) -> InvoicePreparationExportRow:
    sc = scope.sales_contract
    check = decision.amount_check
    return InvoicePreparationExportRow(
        record_type=RECORD_TYPE_SALES_PREPARATION,
        sales_contract_id=str(sc.id),
        sales_contract_no=sc.sales_contract_no,
        our_entity=sc.our_entity,
        customer=sc.customer,
        contract_gross_amount=sc.gross_amount,
        contract_currency=sc.currency,
        linked_procurement_contract_count=len(scope.linked_procurement_contracts),
        confirmed_sales_invoice_count=len(_confirmed_sales_invoices(scope)),
        confirmed_receipt_count=len(_confirmed_in_receipts(scope)),
        comparison_outcome=check.outcome if check else None,
        comparison_message=SALES_COMPARISON_MESSAGES.get(check.outcome) if check else None,
        sales_contract_amount=check.sales_contract_amount if check else None,
        sales_contract_currency=check.sales_contract_currency if check else None,
        declared_amount=check.declared_amount if check else None,
        declared_currency=check.declared_currency if check else None,
        sales_invoice_amount=check.sales_invoice_amount if check else None,
        sales_invoice_currency=check.sales_invoice_currency if check else None,
        # The F1f comparison's resolved trace identifiers, verbatim.
        comparison_shipment_id=str(check.shipment_id) if check and check.shipment_id else None,
        comparison_sales_invoice_id=str(check.sales_invoice_id) if check and check.sales_invoice_id else None,
    )


def _incomplete_association_row(
    record_type: str,
    kind_key: str,
    allocation,
    *,
    scope_id: str,
    scope_no: str,
) -> InvoicePreparationExportRow:
    allocated = getattr(allocation, "allocated_gross_amount", None)
    if allocated is None:
        allocated = getattr(allocation, "allocated_amount", None)
    return InvoicePreparationExportRow(
        record_type=record_type,
        sales_contract_id=scope_id if record_type == RECORD_TYPE_SALES_ATTENTION else None,
        sales_contract_no=scope_no if record_type == RECORD_TYPE_SALES_ATTENTION else None,
        procurement_contract_id=scope_id if record_type == RECORD_TYPE_SUPPLIER_ATTENTION else None,
        contract_no=scope_no if record_type == RECORD_TYPE_SUPPLIER_ATTENTION else None,
        attention_category=ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION,
        attention_code=None,  # no canonical code exists for this context
        attention_message=INCOMPLETE_ASSOCIATION_KIND_LABELS.get(kind_key, kind_key),
        source_id=str(allocation.id),
        allocated_amount=allocated,
    )


def _related_ids_json(ids) -> str | None:
    """Deterministic JSON array of the advisory's full related-id set —
    never truncated to a first id, never a fabricated trace."""
    if not ids:
        return None
    return json.dumps(
        sorted(str(value) for value in ids),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _sales_attention_rows(scope, decision) -> tuple[InvoicePreparationExportRow, ...]:
    sc = scope.sales_contract
    rows: list[InvoicePreparationExportRow] = []
    for work in scope.unresolved_work:
        rows.append(
            InvoicePreparationExportRow(
                record_type=RECORD_TYPE_SALES_ATTENTION,
                sales_contract_id=str(sc.id),
                sales_contract_no=sc.sales_contract_no,
                attention_category=ATTENTION_CATEGORY_UNRESOLVED_WORK,
                attention_code=work.exception_type,
                attention_message=work.summary,
                source_id=str(work.source_id),
            )
        )
    for entry in _incomplete_sales_invoices(scope):
        rows.append(_incomplete_association_row(RECORD_TYPE_SALES_ATTENTION, "SALES_INVOICE", entry.allocation, scope_id=str(sc.id), scope_no=sc.sales_contract_no))
    for entry in _incomplete_sales_receipts(scope):
        rows.append(_incomplete_association_row(RECORD_TYPE_SALES_ATTENTION, "SALES_RECEIPT", entry.allocation, scope_id=str(sc.id), scope_no=sc.sales_contract_no))
    for advisory in decision.advisories:
        rows.append(
            InvoicePreparationExportRow(
                record_type=RECORD_TYPE_SALES_ATTENTION,
                sales_contract_id=str(sc.id),
                sales_contract_no=sc.sales_contract_no,
                attention_category=ATTENTION_CATEGORY_MANAGEMENT_ADVISORY,
                attention_code=advisory.code,
                attention_message=ADVISORY_MESSAGES.get(advisory.code, advisory.code),
                related_invoice_ids=_related_ids_json(getattr(advisory, "related_invoice_ids", ())),
                related_shipment_ids=_related_ids_json(getattr(advisory, "related_shipment_ids", ())),
            )
        )
    return tuple(rows)


def _decimal_json(value: Decimal | None) -> str | None:
    return format(value, "f") if value is not None else None


def _supplier_check_dict(check) -> dict:
    """One check's full canonical trace, preserving the F1 check_name and
    ids — the JSON element never collapses checks into an overall status."""
    if hasattr(check, "compared_invoice_gross_amount"):
        # P02 amount check
        return {
            "kind": "AMOUNT",
            "check_name": check.check_name,
            "outcome": check.outcome,
            "invoice_id": str(check.invoice_id),
            "contract_id": str(check.contract_id),
            "reference_amount": _decimal_json(check.contract_gross_amount),
            "reference_currency": check.contract_currency,
            "invoice_amount": _decimal_json(check.compared_invoice_gross_amount),
            "invoice_currency": check.compared_invoice_currency,
        }
    # P05 item-name check
    return {
        "kind": "ITEM_NAME",
        "check_name": check.check_name,
        "outcome": check.outcome,
        "allocation_id": str(check.allocation_id),
        "contract_item_id": str(check.contract_item_id),
        "invoice_item_id": str(check.invoice_item_id),
        "contract_id": str(check.contract_id),
        "contract_product_name": check.contract_product_name,
        "invoice_product_name": check.invoice_product_name,
    }


def _supplier_checks_json(amount_checks, item_name_checks) -> str | None:
    """Deterministic JSON array of ALL supplier checks — never a fake
    single "overall status". ``sort_keys=True`` keeps the encoding stable;
    the array order is sorted by (kind, check_name, allocation_id) so a
    product name containing ';', '=' or '||' can never corrupt the
    structure (JSON is not parsed by the Summary)."""
    checks = [_supplier_check_dict(c) for c in amount_checks] + [
        _supplier_check_dict(c) for c in item_name_checks
    ]
    if not checks:
        return None
    checks.sort(key=lambda d: (d["kind"], d["check_name"], str(d.get("allocation_id", "") or "")))
    return json.dumps(checks, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _supplier_request_row(scope, decision) -> InvoicePreparationExportRow:
    contract = scope.contract
    return InvoicePreparationExportRow(
        record_type=RECORD_TYPE_SUPPLIER_REQUEST,
        procurement_contract_id=str(contract.id),
        contract_no=contract.contract_no,
        supplier=contract.counterparty,
        our_entity=contract.buyer,
        contract_gross_amount=contract.gross_amount,
        contract_currency=contract.currency,
        expected_invoice_amount=decision.expected_purchase_invoice_gross_amount,
        expected_invoice_currency=decision.expected_purchase_invoice_currency,
        confirmed_purchase_invoice_count=len(_confirmed_purchase_invoices(scope)),
        confirmed_out_payment_count=len(_confirmed_out_payments(scope)),
        supplier_amount_check_outcome=decision.amount_checks[0].outcome if decision.amount_checks else None,
        supplier_amount_check_count=len(decision.amount_checks),
        supplier_item_name_check_count=len(decision.item_name_checks),
        supplier_item_name_deviation_count=sum(
            1 for c in decision.item_name_checks if c.outcome == "DEVIATION"
        ),
        supplier_checks_json=_supplier_checks_json(decision.amount_checks, decision.item_name_checks),
    )


def _supplier_attention_rows(scope, decision) -> tuple[InvoicePreparationExportRow, ...]:
    contract = scope.contract
    rows: list[InvoicePreparationExportRow] = []
    for work in scope.unresolved_work:
        rows.append(
            InvoicePreparationExportRow(
                record_type=RECORD_TYPE_SUPPLIER_ATTENTION,
                procurement_contract_id=str(contract.id),
                contract_no=contract.contract_no,
                attention_category=ATTENTION_CATEGORY_UNRESOLVED_WORK,
                attention_code=work.exception_type,
                attention_message=work.summary,
                source_id=str(work.source_id),
            )
        )
    for entry in _incomplete_purchase_invoices(scope):
        rows.append(_incomplete_association_row(RECORD_TYPE_SUPPLIER_ATTENTION, "PURCHASE_INVOICE", entry.allocation, scope_id=str(contract.id), scope_no=contract.contract_no))
    for entry in _incomplete_out_payments(scope):
        rows.append(_incomplete_association_row(RECORD_TYPE_SUPPLIER_ATTENTION, "OUT_PAYMENT", entry.allocation, scope_id=str(contract.id), scope_no=contract.contract_no))
    for advisory in decision.advisories:
        rows.append(
            InvoicePreparationExportRow(
                record_type=RECORD_TYPE_SUPPLIER_ATTENTION,
                procurement_contract_id=str(contract.id),
                contract_no=contract.contract_no,
                attention_category=ATTENTION_CATEGORY_MANAGEMENT_ADVISORY,
                attention_code=advisory.code,
                attention_message=ADVISORY_MESSAGES.get(advisory.code, advisory.code),
                related_invoice_ids=_related_ids_json(getattr(advisory, "related_invoice_ids", ())),
                related_contract_ids=_related_ids_json(getattr(advisory, "related_contract_ids", ())),
                related_invoice_item_ids=_related_ids_json(getattr(advisory, "related_invoice_item_ids", ())),
            )
        )
    return tuple(rows)


# ---------------------------------------------------------------------------
# Builder — pure projection, no read, no recomputation.
# ---------------------------------------------------------------------------


def build_invoice_preparation_data_product(
    workbench: InvoicePreparationWorkbench,
) -> InvoicePreparationDataProduct:
    """Flatten the Workbench into the neutral Data Product DTO. Accepts
    ONLY the Workbench — never a Session — so the export can never diverge
    from, or re-derive, the F1 rule outputs or re-query repositories."""
    context = workbench.context
    sales_decision_by_id = {d.sales_contract_id: d for d in workbench.sales_report.decisions}
    supplier_decision_by_id = {d.contract_id: d for d in workbench.supplier_report.decisions}

    sales_preparation = tuple(
        _sales_preparation_row(scope, sales_decision_by_id[scope.sales_contract.id])
        for scope in context.sales_scopes
    )
    sales_attention = tuple(
        row
        for scope in context.sales_scopes
        for row in _sales_attention_rows(scope, sales_decision_by_id[scope.sales_contract.id])
    )
    supplier_request = tuple(
        _supplier_request_row(scope, supplier_decision_by_id[scope.contract.id])
        for scope in context.supplier_scopes
    )
    supplier_attention = tuple(
        row
        for scope in context.supplier_scopes
        for row in _supplier_attention_rows(scope, supplier_decision_by_id[scope.contract.id])
    )

    summary = _build_summary(
        sales_preparation=sales_preparation,
        supplier_request=supplier_request,
        sales_attention=sales_attention,
        supplier_attention=supplier_attention,
    )
    return InvoicePreparationDataProduct(
        summary=summary,
        sales_preparation=sales_preparation,
        sales_attention=sales_attention,
        supplier_request=supplier_request,
        supplier_attention=supplier_attention,
    )


def _build_summary(
    *,
    sales_preparation: Iterable[InvoicePreparationExportRow],
    supplier_request: Iterable[InvoicePreparationExportRow],
    sales_attention: Iterable[InvoicePreparationExportRow],
    supplier_attention: Iterable[InvoicePreparationExportRow],
) -> dict[str, int]:
    """Projection-only summary counts — no readiness score, no new
    judgment. Each count is derived from the rows already built from the
    SAME Workbench."""
    sales_prep_rows = list(sales_preparation)
    supplier_req_rows = list(supplier_request)
    sales_attn_rows = list(sales_attention)
    supplier_attn_rows = list(supplier_attention)

    sales_outcomes = Counter(r.comparison_outcome for r in sales_prep_rows if r.comparison_outcome)
    supplier_amount_outcomes = Counter(
        r.supplier_amount_check_outcome for r in supplier_req_rows if r.supplier_amount_check_outcome
    )
    # Counts come from the EXPLICIT neutral fields on the supplier rows —
    # the Summary never parses serialized text (e.g. supplier_checks_json)
    # to recover business structure.
    supplier_amount_check_count = sum((r.supplier_amount_check_count or 0) for r in supplier_req_rows)
    supplier_item_name_check_count = sum((r.supplier_item_name_check_count or 0) for r in supplier_req_rows)
    supplier_item_name_deviation_count = sum(
        (r.supplier_item_name_deviation_count or 0) for r in supplier_req_rows
    )
    attention_categories = Counter(
        r.attention_category for r in sales_attn_rows + supplier_attn_rows if r.attention_category
    )

    summary: dict[str, int] = {
        "sales_scope_count": len(sales_prep_rows),
        "supplier_scope_count": len(supplier_req_rows),
    }
    for outcome in ("MATCH", "DEVIATION", "NOT_COMPARABLE_MISSING_FACT", "NOT_COMPARABLE_CURRENCY_MISMATCH", "NOT_COMPARABLE_AMBIGUOUS_SCOPE"):
        summary[f"sales_comparison_{outcome}"] = sales_outcomes.get(outcome, 0)
    for outcome in ("MATCH", "DEVIATION", "NOT_COMPARABLE_MISSING_FACT", "NOT_COMPARABLE_CURRENCY_MISMATCH"):
        summary[f"supplier_amount_check_{outcome}"] = supplier_amount_outcomes.get(outcome, 0)
    summary["supplier_amount_check_count"] = supplier_amount_check_count
    summary["supplier_item_name_check_count"] = supplier_item_name_check_count
    summary["supplier_item_name_check_DEVIATION"] = supplier_item_name_deviation_count
    for category in (
        ATTENTION_CATEGORY_UNRESOLVED_WORK,
        ATTENTION_CATEGORY_INCOMPLETE_ASSOCIATION,
        ATTENTION_CATEGORY_MANAGEMENT_ADVISORY,
    ):
        summary[f"attention_{category}"] = attention_categories.get(category, 0)
    return summary


# ---------------------------------------------------------------------------
# Serialization — CSV (unified long table) and XLSX (five sheets).
# ---------------------------------------------------------------------------

CSV_HEADERS = [
    "record_type",
    "sales_contract_id",
    "sales_contract_no",
    "procurement_contract_id",
    "contract_no",
    "our_entity",
    "customer",
    "supplier",
    "contract_gross_amount",
    "contract_currency",
    "expected_invoice_amount",
    "expected_invoice_currency",
    "linked_procurement_contract_count",
    "confirmed_sales_invoice_count",
    "confirmed_receipt_count",
    "confirmed_purchase_invoice_count",
    "confirmed_out_payment_count",
    "comparison_outcome",
    "comparison_message",
    "sales_contract_amount",
    "sales_contract_currency",
    "declared_amount",
    "declared_currency",
    "sales_invoice_amount",
    "sales_invoice_currency",
    "comparison_shipment_id",
    "comparison_sales_invoice_id",
    "supplier_amount_check_outcome",
    "supplier_amount_check_count",
    "supplier_item_name_check_count",
    "supplier_item_name_deviation_count",
    "supplier_checks_json",
    "attention_category",
    "attention_code",
    "attention_message",
    "related_invoice_ids",
    "related_contract_ids",
    "related_invoice_item_ids",
    "related_shipment_ids",
    "source_id",
    "allocated_amount",
]


def _fmt_field_value(value: Any) -> str:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None:
        return ""
    return str(value)


def _safe_text(value: Any) -> str:
    """Formula-injection guard, same convention as
    ``period_close_export._safe_text`` / ``contract_ledger_export._safe_text``:
    a leading dangerous character gets a literal-text quote prefix so no
    spreadsheet application ever evaluates exported business text as a
    formula."""
    text = _fmt_field_value(value)
    if text.startswith(_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def _row_to_record(row: InvoicePreparationExportRow) -> list[str]:
    return [_safe_text(getattr(row, col)) for col in CSV_HEADERS]


def export_invoice_preparation_csv(product: InvoicePreparationDataProduct) -> bytes:
    """One unified long-table CSV, every row carrying ``record_type``.
    UTF-8 with BOM (Excel-friendly for Chinese business text), same
    convention as the Period Close / Contract Ledger exports. Deterministic
    row order (the Workbench's own scope order) and stable columns."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
    writer.writerow(CSV_HEADERS)
    for row in product.all_rows:
        writer.writerow(_row_to_record(row))
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def _xlsx_cell(value: Any) -> Any:
    """Decimal amounts stay NUMERIC cells (monetary values usable in a
    spreadsheet); a missing Fact is ``None`` -> blank cell, never 0. Every
    text value passes through the formula-injection guard."""
    if isinstance(value, Decimal):
        return float(value)
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = _fmt_field_value(value)
    if text.startswith(_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _write_rows_sheet(ws, headers: list[str], rows: Iterable[InvoicePreparationExportRow]) -> None:
    _write_header(ws, headers)
    for row in rows:
        ws.append([_xlsx_cell(getattr(row, col)) for col in headers])


_SALES_PREPARATION_COLUMNS = [
    "sales_contract_id",
    "sales_contract_no",
    "our_entity",
    "customer",
    "contract_gross_amount",
    "contract_currency",
    "linked_procurement_contract_count",
    "confirmed_sales_invoice_count",
    "confirmed_receipt_count",
    "comparison_outcome",
    "comparison_message",
    "sales_contract_amount",
    "sales_contract_currency",
    "declared_amount",
    "declared_currency",
    "sales_invoice_amount",
    "sales_invoice_currency",
    "comparison_shipment_id",
    "comparison_sales_invoice_id",
]

_SALES_ATTENTION_COLUMNS = [
    "sales_contract_id",
    "sales_contract_no",
    "attention_category",
    "attention_code",
    "attention_message",
    "related_invoice_ids",
    "related_shipment_ids",
    "source_id",
    "allocated_amount",
]

_SUPPLIER_REQUEST_COLUMNS = [
    "procurement_contract_id",
    "contract_no",
    "supplier",
    "our_entity",
    "contract_gross_amount",
    "contract_currency",
    "expected_invoice_amount",
    "expected_invoice_currency",
    "confirmed_purchase_invoice_count",
    "confirmed_out_payment_count",
    "supplier_amount_check_outcome",
    "supplier_amount_check_count",
    "supplier_item_name_check_count",
    "supplier_item_name_deviation_count",
    "supplier_checks_json",
]

_SUPPLIER_ATTENTION_COLUMNS = [
    "procurement_contract_id",
    "contract_no",
    "attention_category",
    "attention_code",
    "attention_message",
    "related_invoice_ids",
    "related_contract_ids",
    "related_invoice_item_ids",
    "source_id",
    "allocated_amount",
]


def _write_summary_sheet(ws, product: InvoicePreparationDataProduct) -> None:
    _write_header(ws, ["field", "value"])
    for key, value in product.summary.items():
        ws.append([_xlsx_cell(key), _xlsx_cell(value)])


# openpyxl's save_workbook ALWAYS overwrites properties.modified with the
# current time (writer/excel.py), so the docProps/core.xml dcterms:modified
# must be pinned here — the fixed created property survives, modified does
# not. Both are pinned to the same fixed timestamp.
_MODIFIED_TIMESTAMP_RE = re.compile(r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")
_FIXED_XLSX_TIMESTAMP_ISO = "1980-01-01T00:00:00Z"


def _deterministic_xlsx_bytes(content: bytes) -> bytes:
    """openpyxl stamps every zip entry with the CURRENT time and pins
    docProps/core.xml dcterms:modified to the current time at save, so two
    identical exports would differ in raw bytes across a second boundary.
    Rewrite the XLSX zip with fixed entry timestamps AND a fixed core.xml
    modified timestamp so byte identity is reproducible (Phase 2D.3-F2b
    reproducibility requirement). The workbook content is otherwise
    untouched."""
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content), "r") as src, zipfile.ZipFile(
        out, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = _MODIFIED_TIMESTAMP_RE.sub(
                    lambda m: m.group(1) + _FIXED_XLSX_TIMESTAMP_ISO + m.group(2), text
                )
                data = text.encode("utf-8")
            info.date_time = (1980, 1, 1, 0, 0, 0)
            dst.writestr(info, data)
    return out.getvalue()


def export_invoice_preparation_xlsx(product: InvoicePreparationDataProduct) -> bytes:
    """Exactly five sheets, frozen order. Every sheet reads only
    ``InvoicePreparationExportRow`` fields — no independent business
    computation, no raw repository access. Decimal amounts are numeric
    cells; missing Facts are blank. Byte-stable across identical inputs."""
    wb = Workbook()
    # Deterministic workbook properties: openpyxl otherwise stamps the
    # CURRENT time into docProps/core.xml (created/modified), which would
    # break byte identity across a real time boundary. Fixed values here
    # (plus the deterministic zip-entry normalization below) make repeated
    # exports of identical state byte-identical.
    wb.properties.created = _FIXED_XLSX_DATETIME
    wb.properties.modified = _FIXED_XLSX_DATETIME
    _write_summary_sheet(wb.active, product)
    wb.active.title = "01_Summary"
    _write_rows_sheet(wb.create_sheet("02_Sales_Preparation"), _SALES_PREPARATION_COLUMNS, product.sales_preparation)
    _write_rows_sheet(wb.create_sheet("03_Sales_Attention"), _SALES_ATTENTION_COLUMNS, product.sales_attention)
    _write_rows_sheet(wb.create_sheet("04_Supplier_Request"), _SUPPLIER_REQUEST_COLUMNS, product.supplier_request)
    _write_rows_sheet(wb.create_sheet("05_Supplier_Attention"), _SUPPLIER_ATTENTION_COLUMNS, product.supplier_attention)

    buffer = io.BytesIO()
    wb.save(buffer)
    return _deterministic_xlsx_bytes(buffer.getvalue())

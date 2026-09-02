"""Exception & Task Center Data Product (Phase 2D.4-F2).

Turns the F1 read-only Center into the deliverable Exception & Task Data
Product. The ONE application-layer path both Web and CLI call:

    get_unresolved_work_center(session, filters)
        -> build_exception_task_data_product(center)   (this module)
        -> export_exception_task_xlsx() / export_exception_task_csv()

The builder accepts ONLY the F1 ``UnresolvedWorkCenter`` projection — never
a Session, never repositories, never a Period Close rerun, never a rule
evaluation. It is a pure transformation of F1: every row preserves the
frozen neutral fields (source_type/source_id/code/status/summary/created_at/
resolution_route/provenance) plus the full structured scope/id trace, and
the export is exactly the item set the caller's Center filters produced.

The Data Product never adds a source: no UNMATCHED, no R009-R015, no
Invoice Preparation advisory, no MISSING_CONTRACT_GROSS_AMOUNT. It exports
exactly what the F1 Center contains — nothing more, nothing dropped
(unmappable tasks export with blank scope fields, never a guessed
Contract, and never summary parsing).

Determinism and safety reuse the proven techniques of the Period Close /
Invoice Preparation / Contract Ledger exports: fixed workbook properties
and normalized ZIP entry timestamps for byte-identical XLSX, a UTF-8-BOM
QUOTE_ALL CSV, formula-injection neutralization for textual values
(canonical ids/status/code remain exact), and deterministic JSON for
repeatable scope/id arrays (sorted by str(uuid), never first-id truncated).
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from bel.application.unresolved_work_center import (
    ScopeType,
    SourceType,
    UnresolvedWorkCenter,
    UnresolvedWorkItem,
)
from bel.infrastructure.deterministic_xlsx import deterministic_xlsx_bytes, set_fixed_workbook_properties

_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# The canonical machine-readable column set, shared by the unified CSV and
# every source-specific XLSX data sheet. `record_type` (CSV) and
# `source_type` (XLSX) carry the same value — they are the neutral identity
# axis, and the repeatable scope/id fields are always the FULL set as
# deterministic JSON, never a truncated first id.
CSV_HEADERS = [
    "record_type",
    "source_id",
    "code",
    "status",
    "summary",
    "created_at",
    "resolution_route",
    "provenance",
    "procurement_contract_ids",
    "sales_contract_ids",
    "invoice_id",
    "payment_id",
    "shipment_id",
    "match_case_id",
    "scopes_json",
]

XLSX_HEADERS = ["source_type"] + CSV_HEADERS[1:]


@dataclass(frozen=True)
class ExceptionTaskExportRow:
    """One presentation-neutral record. Every Data Product row — CSV or any
    XLSX data sheet — is a projection of this same shape; a field that does
    not apply to this ``record_type`` stays ``None`` (blank cell, never a
    fabricated value). ``source_id`` is the canonical source identity: the
    persisted TaskException/MatchCase id, or the F1 deterministic key for a
    COMPUTED_BLOCKER. Scope/id arrays are pre-serialized deterministic JSON
    strings; a canonical id/status/code stays exact text."""

    record_type: str  # one of SourceType
    source_id: str
    code: str
    status: str
    summary: str
    created_at: str | None
    resolution_route: str
    provenance: str | None
    procurement_contract_ids: str | None
    sales_contract_ids: str | None
    invoice_id: str | None
    payment_id: str | None
    shipment_id: str | None
    match_case_id: str | None
    scopes_json: str | None

    @property
    def source_type(self) -> str:
        """``record_type == source_type`` (docs §5: prefer record_type ==
        source_type) — the XLSX data sheets carry it under its canonical
        name ``source_type``."""
        return self.record_type


@dataclass(frozen=True)
class ExceptionTaskDataProduct:
    """The neutral Data Product DTO. XLSX/CSV serialization reads only this
    — never the Center, never repositories, never the Session."""

    summary: dict[str, Any]
    system_tasks: tuple[ExceptionTaskExportRow, ...]
    match_confirmation: tuple[ExceptionTaskExportRow, ...]
    period_close_blockers: tuple[ExceptionTaskExportRow, ...]

    @property
    def all_rows(self) -> tuple[ExceptionTaskExportRow, ...]:
        return (
            self.system_tasks
            + self.match_confirmation
            + self.period_close_blockers
        )


# ---------------------------------------------------------------------------
# Deterministic JSON for repeatable scope/id arrays
# ---------------------------------------------------------------------------


def _scope_ids_json(item: UnresolvedWorkItem, scope_type: str) -> str | None:
    """The FULL set of ids of one scope kind, as a deterministic JSON array
    sorted by str(uuid) — never truncated to a first id."""
    ids = [scope.scope_id for scope in item.scopes if scope.scope_type == scope_type]
    if not ids:
        return None
    return json.dumps(
        sorted(str(value) for value in ids),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _scopes_json(item: UnresolvedWorkItem) -> str | None:
    """The complete repeatable scope set as deterministic JSON — every scope
    the F1 projection resolved, never only the first candidate. Array order
    is normalized (scope_type, scope_id); dict keys are sorted."""
    if not item.scopes:
        return None
    scopes = [{"scope_type": s.scope_type, "scope_id": str(s.scope_id)} for s in item.scopes]
    scopes.sort(key=lambda d: (d["scope_type"], d["scope_id"]))
    return json.dumps(scopes, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _export_row(item: UnresolvedWorkItem) -> ExceptionTaskExportRow:
    """Pure projection of one F1 Center item — no read, no recomputation."""
    return ExceptionTaskExportRow(
        record_type=item.source_type,
        source_id=str(item.source_id),
        code=item.code,
        status=item.status,
        summary=item.summary,
        created_at=item.created_at.isoformat() if item.created_at is not None else None,
        resolution_route=item.resolution_route,
        provenance=item.provenance,
        procurement_contract_ids=_scope_ids_json(item, ScopeType.PROCUREMENT_CONTRACT),
        sales_contract_ids=_scope_ids_json(item, ScopeType.SALES_CONTRACT),
        invoice_id=str(item.invoice_id) if item.invoice_id is not None else None,
        payment_id=str(item.payment_id) if item.payment_id is not None else None,
        shipment_id=str(item.shipment_id) if item.shipment_id is not None else None,
        match_case_id=str(item.match_case_id) if item.match_case_id is not None else None,
        scopes_json=_scopes_json(item),
    )


def _build_summary(center: UnresolvedWorkCenter) -> dict[str, Any]:
    """Safe export context: counts (from the F1 Center, never recomputed),
    the selected period when supplied, and the non-empty filter context.
    Deterministic key order."""
    summary: dict[str, Any] = {
        "total": center.counts.get("total", 0),
        SourceType.TASK_EXCEPTION: center.counts.get(SourceType.TASK_EXCEPTION, 0),
        SourceType.MATCH_CASE: center.counts.get(SourceType.MATCH_CASE, 0),
        SourceType.COMPUTED_BLOCKER: center.counts.get(SourceType.COMPUTED_BLOCKER, 0),
    }
    filters = center.filters
    if filters.period is not None:
        summary["period"] = filters.period
    if filters.status is not None:
        summary["filter.status"] = filters.status
    if filters.open_only is not None:
        summary["filter.open_only"] = "true" if filters.open_only else "false"
    if filters.source_type is not None:
        summary["filter.source_type"] = filters.source_type
    if filters.code is not None:
        summary["filter.code"] = filters.code
    if filters.procurement_contract_id is not None:
        summary["filter.procurement_contract_id"] = str(filters.procurement_contract_id)
    if filters.sales_contract_id is not None:
        summary["filter.sales_contract_id"] = str(filters.sales_contract_id)
    return summary


# ---------------------------------------------------------------------------
# Builder — pure transformation of the F1 Center projection.
# ---------------------------------------------------------------------------


def build_exception_task_data_product(center: UnresolvedWorkCenter) -> ExceptionTaskDataProduct:
    """Flatten the F1 Center into the neutral Data Product DTO. Accepts
    ONLY the Center projection — never a Session — so the export can never
    diverge from, re-query, or re-derive the F1 item set. The per-source
    row groups preserve the Center's own deterministic item ordering."""
    system_tasks = tuple(_export_row(item) for item in center.items if item.source_type == SourceType.TASK_EXCEPTION)
    match_confirmation = tuple(
        _export_row(item) for item in center.items if item.source_type == SourceType.MATCH_CASE
    )
    period_close_blockers = tuple(
        _export_row(item) for item in center.items if item.source_type == SourceType.COMPUTED_BLOCKER
    )
    return ExceptionTaskDataProduct(
        summary=_build_summary(center),
        system_tasks=system_tasks,
        match_confirmation=match_confirmation,
        period_close_blockers=period_close_blockers,
    )


# ---------------------------------------------------------------------------
# Serialization — CSV (unified long table) and XLSX (four sheets).
# ---------------------------------------------------------------------------


def _fmt_field_value(value: Any) -> str:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None:
        return ""
    return str(value)


def _text_guard(text: str) -> str:
    """Formula-injection guard, same convention as the Period Close /
    Invoice Preparation / Contract Ledger exports: a leading dangerous
    character gets a literal-text quote prefix so no spreadsheet
    application ever evaluates exported business text as a formula.
    Canonical ids/status/code never start with a dangerous prefix, so they
    remain exact text — only genuinely dangerous TEXT values are guarded."""
    if text.startswith(_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def _csv_field(value: Any) -> str:
    """Typed CSV serialization: formula-injection protection applies ONLY
    to TEXT values. None is empty; int stays a plain numeric string. A
    Decimal (not expected in this DTO) is serialized canonically with
    ``format(value, "f")`` and never gains an apostrophe merely because it
    begins with '-'. The neutral DTO here is otherwise all strings."""
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (int, float)):
        return str(value)
    if value is None:
        return ""
    return _text_guard(_fmt_field_value(value))


def _row_to_record(row: ExceptionTaskExportRow) -> list[str]:
    return [_csv_field(getattr(row, col)) for col in CSV_HEADERS]


def export_exception_task_csv(product: ExceptionTaskDataProduct) -> bytes:
    """One unified long-table CSV, every row carrying ``record_type`` (==
    source_type). UTF-8 with BOM (Excel-friendly for Chinese business
    text), same convention as the other Data Products. Deterministic row
    order (the F1 Center's own order) and stable superset columns — every
    XLSX data row is representable here without semantic loss."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
    writer.writerow(CSV_HEADERS)
    for row in product.all_rows:
        writer.writerow(_row_to_record(row))
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def _xlsx_cell(value: Any) -> Any:
    """Typed XLSX serialization: None -> blank cell; int stays typed;
    TEXT values pass through the formula-injection guard. The neutral DTO
    here is Decimal-free, but a Decimal (should one ever appear) follows
    the canonical text policy rather than a blanket Decimal -> float."""
    if isinstance(value, Decimal):
        return _fmt_field_value(value)
    if value is None or isinstance(value, (int, float, bool)):
        return value
    return _text_guard(_fmt_field_value(value))


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _write_rows_sheet(ws, headers: list[str], rows: Iterable[ExceptionTaskExportRow]) -> None:
    _write_header(ws, headers)
    for row in rows:
        ws.append([_xlsx_cell(getattr(row, col)) for col in headers])


def _write_summary_sheet(ws, product: ExceptionTaskDataProduct) -> None:
    _write_header(ws, ["field", "value"])
    for key, value in product.summary.items():
        ws.append([_xlsx_cell(key), _xlsx_cell(value)])


def export_exception_task_xlsx(product: ExceptionTaskDataProduct) -> bytes:
    """Exactly four sheets, frozen order (docs §3):
    01_Summary / 02_System_Tasks / 03_Match_Confirmation /
    04_Period_Close_Blockers. Every sheet reads only
    ``ExceptionTaskExportRow`` fields — no independent business
    computation, no raw repository access. 04 is present but has zero
    rows when no period was requested.

    Byte-stable across identical state and wall-clock time: package
    metadata (ZIP entry timestamps, docProps/core.xml created/modified)
    is pinned through the canonical deterministic-XLSX normalizer."""
    wb = Workbook()
    set_fixed_workbook_properties(wb)
    _write_summary_sheet(wb.active, product)
    wb.active.title = "01_Summary"
    _write_rows_sheet(wb.create_sheet("02_System_Tasks"), XLSX_HEADERS, product.system_tasks)
    _write_rows_sheet(wb.create_sheet("03_Match_Confirmation"), XLSX_HEADERS, product.match_confirmation)
    _write_rows_sheet(wb.create_sheet("04_Period_Close_Blockers"), XLSX_HEADERS, product.period_close_blockers)

    buffer = io.BytesIO()
    wb.save(buffer)
    return deterministic_xlsx_bytes(buffer.getvalue())

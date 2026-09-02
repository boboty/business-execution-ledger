"""Period Close Data Product (Phase 2D.2, docs/PHASE2D2-DECISIONS.md).

Turns the read-only Period Close Workbench into a deliverable,
traceable, reproducible export. The ONE application-layer path both Web
and CLI call:

    get_period_close_workbench(session, period)
        -> build_period_close_data_product()  (this module)
        -> export_period_close_xlsx() / export_period_close_csv()

This module never recomputes a close Decision — it only flattens the
already-composed ``PeriodCloseWorkbench`` (contract/item labels and the
Decision -> Fact -> Evidence trace) into presentation-neutral rows, then
into XLSX/CSV bytes. A blocker is exported exactly as the engine
produced it; no blocker is ever "resolved" here.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from bel.application.period_close_workbench import (
    FactNode,
    PeriodCloseWorkbench,
    WorkbenchAccrual,
    WorkbenchBlocker,
    WorkbenchCandidate,
    WorkbenchDifference,
    WorkbenchReversal,
)
from bel.infrastructure.deterministic_xlsx import deterministic_xlsx_bytes, set_fixed_workbook_properties

RECORD_TYPE_ACCRUAL_REQUIRED = "ACCRUAL_REQUIRED"
RECORD_TYPE_PRIOR_ACCRUAL_REVERSAL = "PRIOR_ACCRUAL_REVERSAL"
RECORD_TYPE_ACTUAL_DIFFERENCE = "ACTUAL_DIFFERENCE"
RECORD_TYPE_CONTRACT_LEVEL_CANDIDATE = "CONTRACT_LEVEL_CANDIDATE"
RECORD_TYPE_BLOCKER = "BLOCKER"

_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


@dataclass(frozen=True)
class PeriodCloseExportRow:
    """One presentation-neutral record. Every Data Product row — CSV or
    any XLSX decision sheet — is a projection of this same shape; a
    field that does not apply to this record_type stays ``None``."""

    record_type: str
    period: str
    period_end: date
    contract_no: str | None = None
    counterparty: str | None = None
    source_item_key: str | None = None
    product_name: str | None = None
    source_period: str | None = None
    quantity: Decimal | None = None
    estimated_cost: Decimal | None = None
    reversal_quantity: Decimal | None = None
    reversal_estimated_cost: Decimal | None = None
    actual_net_cost: Decimal | None = None
    difference: Decimal | None = None
    projected_remaining_quantity: Decimal | None = None
    projected_remaining_cost: Decimal | None = None
    projected_status: str | None = None
    basis: str | None = None
    blocking_reason: str | None = None
    blocker_type: str | None = None
    blocker_context: str | None = None
    evidence_trace: str | None = None


@dataclass(frozen=True)
class PeriodCloseDataProduct:
    """The neutral Data Product DTO. XLSX/CSV serialization reads only
    this — never the Workbench or raw repositories directly."""

    period: str
    period_end: date
    summary: dict[str, int]
    accrual_required: tuple[PeriodCloseExportRow, ...]
    prior_accrual_reversal: tuple[PeriodCloseExportRow, ...]
    actual_difference: tuple[PeriodCloseExportRow, ...]
    contract_level_candidate: tuple[PeriodCloseExportRow, ...]
    blocker: tuple[PeriodCloseExportRow, ...]

    @property
    def all_rows(self) -> tuple[PeriodCloseExportRow, ...]:
        return (
            self.accrual_required
            + self.prior_accrual_reversal
            + self.actual_difference
            + self.contract_level_candidate
            + self.blocker
        )


def _fmt_field_value(value: Any) -> str:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None:
        return ""
    return str(value)


def _fact_node_text(node: FactNode) -> str:
    fields = ";".join(f"{k}={_fmt_field_value(v)}" for k, v in node.fields)
    parts = [f"{node.fact_kind}[{fields}]"]
    if node.document is not None:
        parts.append(f"doc={node.document.file_name}")
    if node.fragment is not None:
        if node.fragment.sheet_name is not None or node.fragment.row_number is not None:
            parts.append(f"loc={node.fragment.sheet_name or ''}:{node.fragment.row_number if node.fragment.row_number is not None else ''}")
        elif node.fragment.locator_json is not None:
            parts.append(f"loc={json.dumps(node.fragment.locator_json, sort_keys=True, ensure_ascii=False)}")
    return "|".join(parts)


def _trace_text(trace: tuple[FactNode, ...]) -> str:
    """A compact, deterministic textual representation of the
    Decision -> Fact -> Evidence chain — never a re-query of the
    database, only a rendering of the trace already composed by
    ``get_period_close_workbench()``."""
    return " -> ".join(_fact_node_text(node) for node in trace)


def _blocker_context_text(context) -> str:
    """Deterministic rendering of the Workbench's existing
    ``BlockerContext`` — current business context explaining the blocker,
    exported as ``blocker_context``. This is NOT Evidence provenance: it
    is never placed in ``evidence_trace``, and it never substitutes for a
    Fact -> Evidence chain (which blockers do not have)."""
    parts: list[str] = []
    if context.historical_source_periods:
        parts.append(f"historical_source_periods={','.join(context.historical_source_periods)}")
    if context.historical_estimated_cost is not None:
        parts.append(f"historical_estimated_cost={_fmt_field_value(context.historical_estimated_cost)}")
    if context.current_remaining_quantity is not None:
        parts.append(f"current_remaining_quantity={_fmt_field_value(context.current_remaining_quantity)}")
    if context.current_remaining_cost is not None:
        parts.append(f"current_remaining_cost={_fmt_field_value(context.current_remaining_cost)}")
    if context.confirmed_invoice_keys:
        parts.append(f"confirmed_invoice_keys={','.join(context.confirmed_invoice_keys)}")
    if context.confirmed_invoice_net_total is not None:
        parts.append(f"confirmed_invoice_net_total={_fmt_field_value(context.confirmed_invoice_net_total)}")
    if context.invoice_item_line_count:
        parts.append(f"invoice_item_line_count={context.invoice_item_line_count}")
    if context.existing_item_allocation_count:
        parts.append(f"existing_item_allocation_count={context.existing_item_allocation_count}")
    if context.cost_recognition_date is not None:
        parts.append(f"cost_recognition_date={context.cost_recognition_date.isoformat()}")
    return ";".join(parts)


def _reversal_row(period: str, period_end: date, row: WorkbenchReversal) -> PeriodCloseExportRow:
    d = row.decision
    return PeriodCloseExportRow(
        record_type=RECORD_TYPE_PRIOR_ACCRUAL_REVERSAL,
        period=period,
        period_end=period_end,
        contract_no=row.contract_no,
        counterparty=row.counterparty,
        source_item_key=row.item.source_item_key if row.item else None,
        product_name=row.item.product_name if row.item else None,
        source_period=d.source_period,
        basis=d.basis,
        reversal_quantity=d.reversal_quantity,
        reversal_estimated_cost=d.reversal_estimated_cost,
        projected_remaining_quantity=d.projected_remaining_quantity,
        projected_remaining_cost=d.projected_remaining_cost,
        projected_status=d.projected_status,
        evidence_trace=_trace_text(row.trace),
    )


def _accrual_row(period: str, period_end: date, row: WorkbenchAccrual) -> PeriodCloseExportRow:
    d = row.decision
    return PeriodCloseExportRow(
        record_type=RECORD_TYPE_ACCRUAL_REQUIRED,
        period=period,
        period_end=period_end,
        contract_no=row.contract_no,
        counterparty=row.counterparty,
        source_item_key=row.item.source_item_key if row.item else None,
        product_name=row.item.product_name if row.item else None,
        quantity=d.quantity,
        estimated_cost=d.estimated_cost,
        basis=d.basis,
        evidence_trace=_trace_text(row.trace),
    )


def _candidate_row(period: str, period_end: date, row: WorkbenchCandidate) -> PeriodCloseExportRow:
    d = row.decision
    return PeriodCloseExportRow(
        record_type=RECORD_TYPE_CONTRACT_LEVEL_CANDIDATE,
        period=period,
        period_end=period_end,
        contract_no=row.contract_no,
        counterparty=row.counterparty,
        estimated_cost=d.estimated_cost,
        blocking_reason=d.blocking_reason,
        evidence_trace=_trace_text(row.trace),
    )


def _difference_row(period: str, period_end: date, row: WorkbenchDifference) -> PeriodCloseExportRow:
    d = row.decision
    return PeriodCloseExportRow(
        record_type=RECORD_TYPE_ACTUAL_DIFFERENCE,
        period=period,
        period_end=period_end,
        contract_no=row.contract_no,
        counterparty=row.counterparty,
        source_item_key=row.item.source_item_key if row.item else None,
        product_name=row.item.product_name if row.item else None,
        actual_net_cost=d.actual_net_cost,
        reversal_estimated_cost=d.reversed_estimated_cost,
        difference=d.difference,
        evidence_trace=_trace_text(row.trace),
    )


def _blocker_row(period: str, period_end: date, row: WorkbenchBlocker) -> PeriodCloseExportRow:
    b = row.blocker
    return PeriodCloseExportRow(
        record_type=RECORD_TYPE_BLOCKER,
        period=period,
        period_end=period_end,
        contract_no=row.contract_no,
        counterparty=row.counterparty,
        source_item_key=row.item.source_item_key if row.item else None,
        product_name=row.item.product_name if row.item else None,
        blocker_type=b.blocker_type,
        # BlockerContext is current business context explaining the
        # blocker — NOT Evidence provenance. It is exported under
        # ``blocker_context`` and never labeled as an evidence trace
        # (no FactNode -> Evidence chain exists for blockers; none is
        # fabricated here).
        blocker_context=_blocker_context_text(row.context),
        evidence_trace=None,
    )


def build_period_close_data_product(workbench: PeriodCloseWorkbench) -> PeriodCloseDataProduct:
    """Flatten the Workbench into the neutral Data Product DTO. Pure
    projection — no read, no recomputation, no new judgment."""
    period = workbench.period
    period_end = workbench.preview.period_end
    return PeriodCloseDataProduct(
        period=period,
        period_end=period_end,
        summary=dict(workbench.summary),
        accrual_required=tuple(_accrual_row(period, period_end, r) for r in workbench.accruals),
        prior_accrual_reversal=tuple(_reversal_row(period, period_end, r) for r in workbench.reversals),
        actual_difference=tuple(_difference_row(period, period_end, r) for r in workbench.differences),
        contract_level_candidate=tuple(_candidate_row(period, period_end, r) for r in workbench.candidates),
        blocker=tuple(_blocker_row(period, period_end, r) for r in workbench.blockers),
    )


CSV_HEADERS = [
    "period",
    "period_end",
    "record_type",
    "contract_no",
    "counterparty",
    "source_item_key",
    "product_name",
    "source_period",
    "quantity",
    "estimated_cost",
    "reversal_quantity",
    "reversal_estimated_cost",
    "actual_net_cost",
    "difference",
    "projected_remaining_quantity",
    "projected_remaining_cost",
    "projected_status",
    "basis",
    "blocking_reason",
    "blocker_type",
    "blocker_context",
    "evidence_trace",
]


def _safe_text(value: Any) -> str:
    """Formula-injection guard, same convention as
    ``contract_ledger_export._safe_text``: a leading dangerous character
    gets a literal-text quote prefix so no spreadsheet application ever
    evaluates exported business text as a formula."""
    text = _fmt_field_value(value)
    if text.startswith(_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def _row_to_record(row: PeriodCloseExportRow) -> list[str]:
    return [
        _safe_text(row.period),
        _safe_text(row.period_end),
        _safe_text(row.record_type),
        _safe_text(row.contract_no),
        _safe_text(row.counterparty),
        _safe_text(row.source_item_key),
        _safe_text(row.product_name),
        _safe_text(row.source_period),
        _safe_text(row.quantity),
        _safe_text(row.estimated_cost),
        _safe_text(row.reversal_quantity),
        _safe_text(row.reversal_estimated_cost),
        _safe_text(row.actual_net_cost),
        _safe_text(row.difference),
        _safe_text(row.projected_remaining_quantity),
        _safe_text(row.projected_remaining_cost),
        _safe_text(row.projected_status),
        _safe_text(row.basis),
        _safe_text(row.blocking_reason),
        _safe_text(row.blocker_type),
        _safe_text(row.blocker_context),
        _safe_text(row.evidence_trace),
    ]


def export_period_close_csv(product: PeriodCloseDataProduct) -> bytes:
    """One unified long-table CSV, every row carrying ``record_type``.
    UTF-8 with BOM (Excel-friendly for Chinese business text), same
    convention as the Contract Business Ledger export."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
    writer.writerow(CSV_HEADERS)
    for row in product.all_rows:
        writer.writerow(_row_to_record(row))
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def _xlsx_cell(value: Any) -> Any:
    if isinstance(value, (Decimal, date, datetime)):
        return _safe_text(value)
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    return "'" + text if text.startswith(_DANGEROUS_PREFIXES) else text


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _write_summary_sheet(ws, product: PeriodCloseDataProduct) -> None:
    _write_header(ws, ["field", "value"])
    ws.append([_xlsx_cell("period"), _xlsx_cell(product.period)])
    ws.append([_xlsx_cell("period_end"), _xlsx_cell(product.period_end)])
    for key in (
        "prior_accrual_reversals",
        "new_accrual_requirements",
        "contract_level_candidates",
        "accrual_actual_differences",
        "blockers",
    ):
        if key in product.summary:
            ws.append([_xlsx_cell(key), _xlsx_cell(product.summary[key])])


_ACCRUAL_REQUIRED_COLUMNS = [
    "period",
    "contract_no",
    "counterparty",
    "source_item_key",
    "product_name",
    "quantity",
    "estimated_cost",
    "basis",
    "evidence_trace",
]

_PRIOR_ACCRUAL_REVERSAL_COLUMNS = [
    "period",
    "contract_no",
    "counterparty",
    "source_item_key",
    "product_name",
    "source_period",
    "basis",
    "reversal_quantity",
    "reversal_estimated_cost",
    "projected_remaining_quantity",
    "projected_remaining_cost",
    "projected_status",
    "evidence_trace",
]

_ACTUAL_DIFFERENCE_COLUMNS = [
    "period",
    "contract_no",
    "counterparty",
    "source_item_key",
    "product_name",
    "actual_net_cost",
    "reversal_estimated_cost",
    "difference",
    "evidence_trace",
]

_CONTRACT_LEVEL_CANDIDATE_COLUMNS = [
    "period",
    "contract_no",
    "counterparty",
    "estimated_cost",
    "blocking_reason",
    "evidence_trace",
]

_BLOCKER_COLUMNS = [
    "period",
    "blocker_type",
    "contract_no",
    "counterparty",
    "source_item_key",
    "product_name",
    "blocker_context",
]


def _write_rows_sheet(ws, headers: list[str], rows: tuple[PeriodCloseExportRow, ...]) -> None:
    _write_header(ws, headers)
    for row in rows:
        ws.append([_xlsx_cell(getattr(row, col)) for col in headers])


def export_period_close_xlsx(product: PeriodCloseDataProduct) -> bytes:
    """Exactly six logical sheets, frozen order (docs/PHASE2D2-DECISIONS.md
    section 4). Every sheet reads only ``PeriodCloseExportRow`` fields —
    no independent business computation, no raw repository access.

    Byte-stable across identical state and wall-clock time: package
    metadata (ZIP entry timestamps, docProps/core.xml created/modified)
    is pinned through the canonical deterministic-XLSX normalizer."""
    wb = Workbook()
    set_fixed_workbook_properties(wb)
    _write_summary_sheet(wb.active, product)
    wb.active.title = "01_Summary"
    _write_rows_sheet(wb.create_sheet("02_Accrual_Required"), _ACCRUAL_REQUIRED_COLUMNS, product.accrual_required)
    _write_rows_sheet(
        wb.create_sheet("03_Prior_Accrual_Reversal"), _PRIOR_ACCRUAL_REVERSAL_COLUMNS, product.prior_accrual_reversal
    )
    _write_rows_sheet(wb.create_sheet("04_Actual_Difference"), _ACTUAL_DIFFERENCE_COLUMNS, product.actual_difference)
    _write_rows_sheet(
        wb.create_sheet("05_Contract_Level_Candidate"),
        _CONTRACT_LEVEL_CANDIDATE_COLUMNS,
        product.contract_level_candidate,
    )
    _write_rows_sheet(wb.create_sheet("06_Blockers"), _BLOCKER_COLUMNS, product.blocker)

    buffer = io.BytesIO()
    wb.save(buffer)
    return deterministic_xlsx_bytes(buffer.getvalue())

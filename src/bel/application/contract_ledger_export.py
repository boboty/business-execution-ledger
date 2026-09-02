"""Contract Business Ledger Data Product — CSV and XLSX export (Phase
2D.1-R4, docs/ROADMAP.md sections 27-30).

Both formats flatten the SAME ``ContractBusinessLedger`` projection the
web page renders (docs section 26) — there is no independent query here,
only presentation-neutral-to-spreadsheet serialization. Repeated
dimensions (items, shipments, linked sales scopes, allocations) are
encoded as deterministic JSON arrays in dedicated columns on the CSV's
one row per procurement contract, and as separate detail sheets on the
XLSX linked back by ``procurement_contract_id`` / a scope id.

Every business string that reaches a cell goes through
``_safe_text`` / ``_xlsx_cell`` first — a formula-injection guard (section 30, HARD
security requirement): a value beginning with ``=``, ``+``, ``-``, ``@``,
a tab, or a carriage return is neutralized so no spreadsheet application
ever interprets Ledger content (customer names, supplier names, product
names, contract numbers — all ultimately sourced from imported Evidence)
as a formula.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from bel.application.contract_business_ledger import ContractBusinessLedger, ContractLedgerRow
from bel.infrastructure.deterministic_xlsx import deterministic_xlsx_bytes, set_fixed_workbook_properties

_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_text(value: Any) -> str:
    """Render any value as a formula-injection-safe plain string. A
    leading dangerous character gets a single quote prefix — Excel/Sheets
    render a leading ``'`` as a literal-text marker and never evaluate
    what follows, and CSV consumers simply see the literal quote plus the
    original text, never a computed value."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        text = format(value, "f")
    elif isinstance(value, (date, datetime)):
        text = value.isoformat()
    else:
        text = str(value)
    if text.startswith(_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def _safe_json(value: Any) -> str:
    """Same neutralization applied to a JSON-serialized structured cell
    (section 29) — the JSON text itself must not be interpretable as a
    formula by a spreadsheet application that opens the CSV."""
    text = json.dumps(value, ensure_ascii=False, sort_keys=False)
    if text.startswith(_DANGEROUS_PREFIXES):
        return "'" + text
    return text


def _decimal_json(value: Decimal | None) -> str | None:
    return format(value, "f") if value is not None else None


def _item_json(item) -> dict:
    return {
        "source_item_key": item.source_item_key,
        "sku": item.sku,
        "product_name": item.product_name,
        "specification": item.specification,
        "quantity": _decimal_json(item.quantity),
        "unit": item.unit,
        "unit_price": _decimal_json(item.unit_price),
        "gross_amount": _decimal_json(item.gross_amount),
        "net_amount": _decimal_json(item.net_amount),
    }


def _shipment_json(entry) -> dict:
    s = entry.shipment
    return {
        "shipment_id": str(s.id),
        "external_reference": s.external_reference,
        "execution_date": s.execution_date.isoformat() if s.execution_date else None,
        "contract_item_id": str(s.contract_item_id) if s.contract_item_id else None,
        "quantity": _decimal_json(s.quantity),
    }


def _procurement_invoice_json(entry) -> dict:
    return {
        "invoice_external_key": entry.invoice.external_invoice_key if entry.invoice else None,
        "allocated_gross_amount": _decimal_json(entry.allocation.allocated_gross_amount),
        "confirmation_type": entry.allocation.confirmation_type,
    }


def _outgoing_payment_json(entry) -> dict:
    return {
        "bank_reference": entry.payment.bank_reference if entry.payment else None,
        "allocated_amount": _decimal_json(entry.allocation.allocated_amount),
        "confirmation_type": entry.allocation.confirmation_type,
    }


def _accrual_json(entry) -> dict:
    return {
        "accrual_id": str(entry.accrual.id),
        "contract_item_id": str(entry.contract_item_id),
        "period": entry.accrual.period,
        "remaining_quantity": _decimal_json(entry.remaining_quantity),
        "remaining_estimated_cost": _decimal_json(entry.remaining_estimated_cost),
        "reversed_quantity": _decimal_json(entry.reversed_quantity),
        "reversed_estimated_cost": _decimal_json(entry.reversed_estimated_cost),
        "current_status": entry.projected_status,
    }


def _sales_scope_json(scope) -> dict:
    """Scope-level facts only — NEVER a figure attributed to the
    procurement contract that links to it (section 13/28's HARD rule)."""
    return {
        "sales_contract_id": str(scope.sales_contract.id),
        "sales_contract_no": scope.sales_contract.sales_contract_no,
        "our_entity": scope.sales_contract.our_entity,
        "customer": scope.sales_contract.customer,
        "currency": scope.sales_contract.currency,
        "gross_amount": _decimal_json(scope.sales_contract.gross_amount),
        "contract_date": scope.sales_contract.contract_date.isoformat() if scope.sales_contract.contract_date else None,
        "sales_invoice_confirmed_allocations": [
            {
                "invoice_external_key": a.invoice.external_invoice_key if a.invoice else None,
                "allocated_gross_amount": _decimal_json(a.allocation.allocated_gross_amount),
            }
            for a in scope.sales_invoice_allocations
        ],
        "incoming_receipt_confirmed_allocations": [
            {
                "bank_reference": a.payment.bank_reference if a.payment else None,
                "allocated_amount": _decimal_json(a.allocation.allocated_amount),
            }
            for a in scope.incoming_receipt_allocations
        ],
        "has_unresolved": scope.has_unresolved,
    }


def _unresolved_json(row: ContractLedgerRow) -> list[dict]:
    return [
        {"source": w.source, "exception_type": w.exception_type, "summary": w.summary, "source_id": str(w.source_id)}
        for w in row.unresolved_work
    ]


CSV_HEADERS = [
    "contract_id",
    "contract_no",
    "supplier_counterparty",
    "our_entity_buyer",
    "gross_amount",
    "currency",
    "contract_date",
    "items_json",
    "shipments_json",
    "procurement_invoices_json",
    "outgoing_payments_json",
    "accruals_json",
    "linked_sales_scopes_json",
    "unresolved_work_json",
    "has_unresolved",
    "outbound_invoice_preparation_state",
]

OUTBOUND_INVOICE_STATE_NOT_EVALUATED = "NOT_EVALUATED_BY_RULE"


def _row_to_record(row: ContractLedgerRow) -> list[str]:
    c = row.contract
    return [
        _safe_text(c.id),
        _safe_text(c.contract_no),
        _safe_text(c.counterparty),
        _safe_text(c.buyer),
        _safe_text(c.gross_amount),
        _safe_text(c.currency),
        _safe_text(c.contract_date),
        _safe_json([_item_json(i) for i in row.items]),
        _safe_json([_shipment_json(s) for s in row.shipments]),
        _safe_json([_procurement_invoice_json(i) for i in row.procurement_invoices]),
        _safe_json([_outgoing_payment_json(p) for p in row.outgoing_payments]),
        _safe_json([_accrual_json(a) for a in row.accruals]),
        _safe_json([_sales_scope_json(s) for s in row.sales_scopes]),
        _safe_json(_unresolved_json(row)),
        _safe_text(row.has_unresolved),
        # Phase 2D.3's eligibility rule is not frozen (docs/PHASE2D1-R0-DECISIONS.md
        # section 3.6) — this column is capability metadata, never a
        # business judgment. It is always this one constant.
        OUTBOUND_INVOICE_STATE_NOT_EVALUATED,
    ]


def export_contract_business_ledger_csv(ledger: ContractBusinessLedger) -> bytes:
    """UTF-8 with BOM (Excel-friendly for Chinese business text). One row
    per procurement contract — repeated dimensions are lossless
    deterministic JSON cells, never row explosion (section 29)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
    writer.writerow(CSV_HEADERS)
    for row in ledger.rows:
        writer.writerow(_row_to_record(row))
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def export_contract_business_ledger_xlsx(ledger: ContractBusinessLedger) -> bytes:
    """Sheet 1 is the primary Ledger (one row per procurement contract,
    same flattening as CSV). Detail sheets are optional lossless
    breakouts, each explicitly keyed by ``procurement_contract_id`` (and,
    for the sales-scope sheet, ``sales_contract_id``) — never an
    apportioned amount (section 28).

    Byte-stable across identical state and wall-clock time: package
    metadata (ZIP entry timestamps, docProps/core.xml created/modified)
    is pinned through the canonical deterministic-XLSX normalizer."""
    wb = Workbook()
    set_fixed_workbook_properties(wb)
    _write_main_sheet(wb.active, ledger)
    wb.active.title = "Contract Business Ledger"
    _write_items_sheet(wb.create_sheet("Contract Items"), ledger)
    _write_shipments_sheet(wb.create_sheet("Shipments"), ledger)
    _write_sales_scopes_sheet(wb.create_sheet("Linked Sales Scopes"), ledger)

    buffer = io.BytesIO()
    wb.save(buffer)
    return deterministic_xlsx_bytes(buffer.getvalue())


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _xlsx_cell(value: Any) -> Any:
    """openpyxl already writes plain str values as literal text, never as
    a formula — a cell only becomes a formula if the value is passed
    with ``data_type='f'`` or the string is later re-interpreted by a
    consumer that DOES treat a leading '=' as a trigger (e.g. Excel
    itself, on open). Neutralizing the leading character defensively
    here, exactly as the CSV path does, is the same HARD requirement
    (section 30) applied to the other artifact."""
    if isinstance(value, (Decimal, date, datetime)):
        return _safe_text(value)
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    text = str(value)
    return "'" + text if text.startswith(_DANGEROUS_PREFIXES) else text


def _write_main_sheet(ws, ledger: ContractBusinessLedger) -> None:
    _write_header(ws, CSV_HEADERS)
    for row in ledger.rows:
        record = _row_to_record(row)
        # _row_to_record already ran every field through the safe-text /
        # safe-json guards; write as literal strings.
        ws.append(record)


def _write_items_sheet(ws, ledger: ContractBusinessLedger) -> None:
    headers = [
        "procurement_contract_id",
        "contract_no",
        "source_item_key",
        "sku",
        "product_name",
        "specification",
        "quantity",
        "unit",
        "unit_price",
        "gross_amount",
        "net_amount",
    ]
    _write_header(ws, headers)
    for row in ledger.rows:
        for item in row.items:
            ws.append(
                [
                    _xlsx_cell(row.contract.id),
                    _xlsx_cell(row.contract.contract_no),
                    _xlsx_cell(item.source_item_key),
                    _xlsx_cell(item.sku),
                    _xlsx_cell(item.product_name),
                    _xlsx_cell(item.specification),
                    _xlsx_cell(item.quantity),
                    _xlsx_cell(item.unit),
                    _xlsx_cell(item.unit_price),
                    _xlsx_cell(item.gross_amount),
                    _xlsx_cell(item.net_amount),
                ]
            )


def _write_shipments_sheet(ws, ledger: ContractBusinessLedger) -> None:
    headers = [
        "procurement_contract_id",
        "contract_no",
        "shipment_id",
        "external_reference",
        "execution_date",
        "contract_item_id",
        "quantity",
    ]
    _write_header(ws, headers)
    for row in ledger.rows:
        for entry in row.shipments:
            s = entry.shipment
            ws.append(
                [
                    _xlsx_cell(row.contract.id),
                    _xlsx_cell(row.contract.contract_no),
                    _xlsx_cell(s.id),
                    _xlsx_cell(s.external_reference),
                    _xlsx_cell(s.execution_date),
                    _xlsx_cell(s.contract_item_id),
                    _xlsx_cell(s.quantity),
                ]
            )


def _write_sales_scopes_sheet(ws, ledger: ContractBusinessLedger) -> None:
    """Scope-level facts only, per procurement row that links to them —
    the SAME scope may legitimately appear on more than one procurement
    contract's rows here (spec section 13/28's explicit requirement).
    Never an apportioned figure."""
    headers = [
        "procurement_contract_id",
        "contract_no",
        "sales_contract_id",
        "sales_contract_no",
        "our_entity",
        "customer",
        "currency",
        "gross_amount",
        "contract_date",
        "sales_invoice_confirmed_allocation_count",
        "sales_invoice_confirmed_allocation_total",
        "incoming_receipt_confirmed_allocation_count",
        "incoming_receipt_confirmed_allocation_total",
        "has_unresolved",
    ]
    _write_header(ws, headers)
    for row in ledger.rows:
        for scope in row.sales_scopes:
            invoice_total = sum(
                (a.allocation.allocated_gross_amount for a in scope.sales_invoice_allocations), Decimal("0")
            )
            receipt_total = sum(
                (a.allocation.allocated_amount for a in scope.incoming_receipt_allocations), Decimal("0")
            )
            ws.append(
                [
                    _xlsx_cell(row.contract.id),
                    _xlsx_cell(row.contract.contract_no),
                    _xlsx_cell(scope.sales_contract.id),
                    _xlsx_cell(scope.sales_contract.sales_contract_no),
                    _xlsx_cell(scope.sales_contract.our_entity),
                    _xlsx_cell(scope.sales_contract.customer),
                    _xlsx_cell(scope.sales_contract.currency),
                    _xlsx_cell(scope.sales_contract.gross_amount),
                    _xlsx_cell(scope.sales_contract.contract_date),
                    len(scope.sales_invoice_allocations),
                    _xlsx_cell(invoice_total),
                    len(scope.incoming_receipt_allocations),
                    _xlsx_cell(receipt_total),
                    scope.has_unresolved,
                ]
            )
